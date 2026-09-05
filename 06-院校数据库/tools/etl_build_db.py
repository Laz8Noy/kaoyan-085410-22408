# -*- coding: utf-8 -*-
"""考研院校数据库 AI 版 v1 构建器（ETL）。
输入：终极版 20260826 xlsx（桌面）+ 王道考情链接总表 xlsx + kaoqing json + N诺(内含于终极版)。
输出：deliverables/20260903-考研院校数据库AI版/data/schools/*.json + data/wangdao_links_all.json
"""
import os, re, json, sys
from openpyxl import load_workbook

WS = r"<SOURCE_DIR>"
ZB = os.path.join(WS, r"06_终极版输出\全国408_085410双非热度版_终极版_20260826.xlsx")
WD = os.path.join(WS, r"deliverables\20260825-王道考情链接总表\王道考情链接总表_20260825.xlsx")
KQ = r"E:\some work\_归档\202608_408中间数据\kaoqing_20260821.json"
OUTDIR = os.path.join(WS, r"deliverables\20260903-考研院校数据库AI版")
SCH = os.path.join(OUTDIR, "data", "schools")

def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("—", "-", "", "None", "nan"):
        return None
    return s

def sheet_rows(path, sheet, name_idx=0):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return [], []
    header = [norm(c) or f"__c{i}" for i, c in enumerate(rows[name_idx])]
    data = []
    for r in rows[name_idx + 1:]:
        d = {}
        for i, c in enumerate(r):
            if i < len(header):
                d[header[i]] = norm(c)
        data.append(d)
    return header, data

def parse_lines(v):
    """'— / — / 260 / 264' -> {'2023':null,'2024':null,'2025':260,'2026':264}；单个值视为2026。"""
    out = {"2023": None, "2024": None, "2025": None, "2026": None}
    if not v:
        return out
    parts = [p.strip() for p in str(v).replace("，", ",").split("/")]
    if len(parts) == 1 and parts[0]:
        out["2026"] = parts[0]
    else:
        ys = ["2023", "2024", "2025", "2026"]
        for y, p in zip(ys, parts):
            out[y] = norm(p)
    return out

def load_sheet(header, data, colmap):
    """返回按 colmap(header名->字段) 转换的行 dict 列表。"""
    out = []
    for d in data:
        rec = {}
        for h, f in colmap.items():
            rec[f] = d.get(h)
        out.append(rec)
    return out

print("[1] 读取终极版各 Sheet ...")
_, zb_total = sheet_rows(ZB, "总览")
_, zb_region = sheet_rows(ZB, "区域榜")
_, zb_recruit = sheet_rows(ZB, "招录比对比")
_, zb_src = sheet_rows(ZB, "三源核对")
_, zb_kq = sheet_rows(ZB, "考情明细")
_, zb_nn = sheet_rows(ZB, "N诺新增候选校")
_, zb_tj = sheet_rows(ZB, "调剂院校")
_, zb_zd = sheet_rows(ZB, "重点院校")
_, zb_2027 = sheet_rows(ZB, "2027改考动态")
_, zb_tutor = sheet_rows(ZB, "导师研究方向")
_, zb_dsrc = sheet_rows(ZB, "数据源")
_, zb_hot = sheet_rows(ZB, "热度榜")

def pick(d, keys):
    for k in keys:
        if d.get(k) is not None:
            return d.get(k)
    return None

units = {}   # key=(校,学院) -> dict

def get_unit(school, college):
    k = (school, college or "（全校）")
    if k not in units:
        units[k] = {"school": school, "college": college or "（全校）", "srcRows": []}
    return units[k], k

def merge_unit(k, d, src):
    u = units[k]
    for kk, vv in d.items():
        if vv is not None:
            if kk not in u or u[kk] is None:
                u[kk] = vv
    u.setdefault("src", []).append(src)

# ---- 总览（主体行，字段最全）----
for r in zb_total:
    school, college = r.get("院校"), r.get("学院")
    if not school:
        continue
    u, k = get_unit(school, college)
    merge_unit(k, {
        "category": "S-主体",
        "province": r.get("省份"), "region": r.get("大区"), "tier": r.get("层次"),
        "subjectClass": r.get("初试科目"), "direction": r.get("方向"),
        "line2026": r.get("2026复试线"), "lineDelta": r.get("线差"),
        "linesByYear": parse_lines(r.get("历年线(23/24/25/26)")),
        "plan2026": r.get("拟招"), "fill": r.get("一志愿/调剂"),
        "ratioApply": r.get("报录比"), "ratioRetest": r.get("复录比"),
        "retestCnt": r.get("复试人数"), "admitCnt": r.get("录取人数"),
        "admitMax": r.get("录取最高"), "admitMin": r.get("录取最低"), "admitAvg": r.get("录取平均"),
        "scope": r.get("口径"), "aiTag": r.get("AI方向备注"),
        "heatNet": r.get("网络热度"), "heatComp": r.get("综合热度"),
        "nn408avg": r.get("408均分(N诺)"), "nnRate": r.get("录取率(N诺)"),
        "note": r.get("备注"), "srcLabel": r.get("来源"),
    }, "总览")

# ---- 区域榜：补齐对照校 C + 热度 ----
for r in zb_region:
    school = r.get("院校")
    if not school:
        continue
    college = r.get("学院")
    cat = r.get("类别")
    u, k = get_unit(school, college)
    if cat and cat != "主体":
        u["category"] = "C-对照" if cat == "对照" else cat
    merge_unit(k, {
        "province": r.get("省份"), "region": r.get("大区"), "tier": r.get("层次"),
        "subjectClass": r.get("初试科目"), "line2026": r.get("2026线"),
        "lineDelta": r.get("线差"), "direction": r.get("方向"),
        "heatComp": r.get("综合热度"), "heatNet": r.get("网络热度"), "aiTag": r.get("AI方向备注"),
        "note": r.get("备注"), "srcLabel": r.get("来源"),
    }, "区域榜")

# ---- 招录比对比：补报录/复录/热度 ----
for r in zb_recruit:
    school, college = r.get("院校"), r.get("学院")
    if not school:
        continue
    u, k = get_unit(school, college)
    merge_unit(k, {
        "region": r.get("大区"), "tier": r.get("层次"),
        "subjectClass": r.get("初试科目"), "line2026": r.get("2026线"), "lineDelta": r.get("线差"),
        "plan2026": r.get("拟招"), "fill": r.get("一志愿/调剂"),
        "ratioApply": r.get("报录比"), "ratioRetest": r.get("复录比"),
        "retestCnt": r.get("复试人数"), "admitCnt": r.get("录取人数"),
        "admitMax": r.get("录取最高"), "admitMin": r.get("录取最低"), "admitAvg": r.get("录取平均"),
        "heatComp": r.get("竞争热度"), "srcLabel": r.get("来源"),
    }, "招录比对比")

# ---- 三源核对：NN + 川渝 + 王道 + 标注（S 与 C 都覆盖；表无学院列，尽量按 N诺院系映射到已有 unit）----
def best_unit_for(school, nn_college):
    """按 nnCollege 与既有单位学院名做包含匹配；单单位时直接返回；找不到返回 None。"""
    lst = []
    for (s2, c2) in units:
        if s2 == school and c2 and c2 != "（全校）":
            lst.append((c2, (s2, c2)))
    if not lst:
        return None
    if nn_college:
        for c2, key in lst:
            if nn_college in c2 or c2 in nn_college:
                return key
    if len(lst) == 1:
        return lst[0][1]
    return None

cross_refs = {}
crs_seen = {}
for r in zb_src:
    school = r.get("院校")
    if not school:
        continue
    nn_college = r.get("N诺院系")
    tk = best_unit_for(school, nn_college)
    payload = {
        "tier": r.get("层次"), "lineUltimate2026": r.get("终极版2026线"),
        "nnProg": r.get("N诺专业"), "nnCollege": nn_college,
        "nnSubjects": r.get("N诺初试"), "nn408avg": r.get("N诺408均分"), "nnRate": r.get("N诺录取率"),
        "cyCode": r.get("川渝专业代码"), "cyRetest": r.get("川渝复试人数"),
        "cyAdmit": r.get("川渝录取人数"), "cyAvg": r.get("川渝均分"),
        "wdCount": r.get("王道链接数"), "wdYears": r.get("王道覆盖年份"),
        "wdUrl26": r.get("王道26链接"), "remark": r.get("标注"),
    }
    if tk is not None:
        u = units[tk]
        if not u.get("category"):
            u["category"] = "C-对照" if r.get("类别") == "对照" else "S-主体"
        merge_unit(tk, payload, "三源核对")
    else:
        key = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        seen = crs_seen.setdefault(school, set())
        if key not in seen:
            cross_refs.setdefault(school, []).append(payload)
            seen.add(key)

# ---- 考情明细（官方细节行）----
kq_detail = {}
for r in zb_kq:
    school, college = r.get("院校"), r.get("学院")
    if not school:
        continue
    u, k = get_unit(school, college)
    rec = {
        "college": college, "tier": r.get("层次"), "subjects": r.get("初试科目"),
        "line2026": r.get("2026线"), "plan": r.get("拟招"),
        "retestCnt": r.get("复试人数"), "admitCnt": r.get("录取人数"),
        "ratioRetest": r.get("复录比"), "ratioApply": r.get("报录比"),
        "admitMax": r.get("录取最高"), "admitMin": r.get("录取最低"), "admitAvg": r.get("录取平均"),
        "scope": r.get("口径"), "verify": r.get("核实状态"), "url": r.get("来源URL"),
    }
    kq_detail[k] = rec
    merge_unit(k, {"kaoqingUrl": rec["url"], "scope": rec["scope"] if not u.get("scope") else u["scope"]}, "考情明细")

# ---- N诺新增候选校（独立候补清单）----
nn_cand = []
for r in zb_nn:
    school = r.get("院校")
    if not school:
        continue
    nn_cand.append({
        "school": school, "province": r.get("省份"), "college": r.get("院系"),
        "program": r.get("专业"), "subjects": r.get("初试"), "line": r.get("复试线"),
        "retestCnt": r.get("复试人数"), "adjCnt": r.get("调剂人数"),
        "avgTotal": r.get("复试总分均分"), "avg408": r.get("408均分"),
        "avgPol": r.get("政治均分"), "avgEng": r.get("英语均分"), "avgMath": r.get("数学均分"),
        "admitCnt": r.get("录取人数"), "admitAvg": r.get("录取均分"),
        "rate": r.get("录取率"), "ratioNN": r.get("报录比(N诺=复试/录取)"), "note": r.get("备注"),
    })

# ---- 调剂院校 / 重点院校 ----
tiaji = {}
for r in zb_tj:
    s = r.get("院校")
    if s:
        tiaji[s] = {"subjects": r.get("初试科目"), "region": r.get("大区"), "tier": r.get("层次"),
                    "line": r.get("2026线"), "fill": r.get("一志愿/调剂"),
                    "aiTag": r.get("AI方向"), "src": r.get("来源")}
zd = {}
for r in zb_zd:
    s = r.get("院校")
    if s:
        zd[s] = dict(r)

# ---- 2027 改考动态 / 导师 ----
up2027 = {}
for r in zb_2027:
    s = r.get("院校")
    if s:
        up2027.setdefault(s, []).append(dict(r))
tutors = {}
for r in zb_tutor:
    s = r.get("院校")
    if s:
        tutors.setdefault(s, []).append(r.get("导师及研究方向（含Agent/大模型/具身智能关键词）"))

# ---- 数据源：按院校名关联 ----
dsrc = {}
for r in zb_dsrc:
    label = r.get("来源") or r.get("__c1")
    url = r.get("URL") or r.get("__c2")
    if label and url:
        dsrc[label] = url
def collect_sources(school):
    got = []
    for label, url in dsrc.items():
        if school.replace("（", "(").replace("）", ")") in label.replace("（", "(").replace("）", ")"):
            got.append({"label": label, "url": url})
    return got

# ---- 王道全量链接（链接明细：序号/省份/院校/层次/年份/URL，位置列解析）----
print("[2] 读取王道链接总表 ...")
wd_by_school = {}
wb = load_workbook(WD, read_only=True, data_only=True)
ws = wb["链接明细"]
for row in ws.iter_rows(values_only=True):
    vals = [norm(c) for c in row]
    if len(vals) >= 6 and vals[1] and vals[2] and vals[4] and vals[5]:
        school, year, url = vals[2], vals[4], vals[5]
        wd_by_school.setdefault(school, {})[str(year)] = url
wb.close()

# ---- kaoqing json（2026 085410 细节，学校|学院 key）----
kqj = {}
if os.path.exists(KQ):
    with open(KQ, "r", encoding="utf-8") as f:
        kqj = json.load(f).get("schools", {})

# ---- 汇总 school 文件 ----
print("[3] 生成 school JSON ...")
schools = {}
for (school, college), u in units.items():
    schools.setdefault(school, []).append((college, u))

ids = {}
def sid(name):
    base = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", name)
    if base not in ids:
        ids[base] = 0
    ids[base] += 1
    return f"{ids[base]:03d}-{base}"

os.makedirs(SCH, exist_ok=True)
meta_schools = []

for school, unit_list in sorted(schools.items()):
    # 从第一个 unit 取基本信息
    uni0 = unit_list[0][1]
    units_out = []
    for college, u in sorted(unit_list):
        rec = {k: v for k, v in u.items() if k not in ("school", "college", "srcRows")}
        rec["college"] = college
        units_out.append(rec)
    # N诺候选（若本校本不在主体行）
    nn_here = [c for c in nn_cand if c["school"] == school]
    kq_rows = [rec for (s2, c2), rec in kq_detail.items() if s2 == school]
    xrefs = cross_refs.get(school, [])
    obj = {
        "schema": "kaoyan-school/v1",
        "name": school,
        "units": units_out,
        "crossRefs": xrefs,
        "kaoqingDetail2026": kq_rows,
        "nnCandidates": nn_here,
        "wangdaoLinks": wd_by_school.get(school, {}),
        "updates2027": up2027.get(school, []),
        "tutors": tutors.get(school, []),
        "inTJList": tiaji.get(school),
        "inKeyList": zd.get(school),
        "conflicts": [],
        "sources": collect_sources(school),
        "generated": {"by": "etl_build_db.py", "date": "2026-09-03",
                      "src": ["全国408_085410双非热度版_终极版_20260826.xlsx",
                              "王道考情链接总表_20260825.xlsx", "kaoqing_20260821.json"],
                      "caution": "机器自动整合，主体行口径=085410 相关招生；数字为资料时点(2026)汇总，正式报考以当年官方简章/研招网为准。"},
    }
    # 显式冲突：科目分类 vs N诺初试
    for un in units_out:
        if un.get("subjectClass") and un.get("nnSubjects"):
            a, b = un["subjectClass"], un["nnSubjects"]
            if a != b and a in ("22408", "混合") and ("数一" in str(b) or "11408" in str(b) or "数一408" in str(b)):
                obj["conflicts"].append({
                    "field": "初试科目口径",
                    "unit": un["college"],
                    "claims": [{"src": "终极版0626-初试科目列", "value": a},
                               {"src": "N诺-初试", "value": b}],
                    "status": "待核", "action": "以学校2026/2027招生目录原文为准",
                })
    for cr in xrefs:
        if cr.get("remark") and "参考085411" in str(cr["remark"]):
            obj["conflicts"].append({
                "field": "考情参考口径", "unit": "跨学院参考行",
                "claims": [{"src": "三源核对-标注", "value": cr["remark"]},
                           {"src": "N诺", "value": f"{cr.get('nnProg')} {cr.get('nnSubjects')} 408均{cr.get('nn408avg')} 录取率{cr.get('nnRate')}"}],
                "status": "参考", "action": "该学院 085410 无独立N诺数据，用 085411 大数据(英二数二408)近似；085410 实际口径以官方为准",
            })
    if kqj.get(f"{school}|{college}") and False:
        pass

    # CUIT 深档补丁（来自 20260903-成都信息工程大学专档 人工核对结论）
    if school == "成都信息工程大学":
        obj["deepRef"] = "deliverables/20260903-成都信息工程大学专档/成都信息工程大学_考研信息全档_20260903.md"
        obj["note"] = ("085410 为专业代码(人工智能专硕)，≠考试科目。两个学院均招 085410：计算机学院(计划约25/一志愿2复1录)与人工智能学院(019,计划12含推2/一志愿1人271)。"
                       "2026 为改 408 首年；AI 学院官方目录 085410=英二+数一+408；计院 085410/085404 数学一/数二两派资料冲突；doubao 深挖称 2027 专硕改数二(未官宣)。"
                       "若 2027 改数二则真正匹配 22408；085411 大数据(软工院/应用数学院)为 2026 唯一确认英二数二408 的专业。")
        for un in obj["units"]:
            if un["college"] == "计算机学院":
                un["subjectClaim2026"] = [
                    {"claim": "英二+数学一+408", "src": "N诺2026表 / doubao深挖横评"},
                    {"claim": "英二+数学二+408", "src": "王道OCR 26考情 img10(计院)"},
                ]
                un["subjectStatus"] = "待核"
            if un["college"] == "人工智能学院":
                un["subjectClaim2026"] = [{"claim": "英二+数学一+408", "src": "官方2026招生目录OCR(王道img08) / doubao / N诺"}]
                un["subjectStatus"] = "官方口径(2026)"
        obj["conflicts"] += [
            {"id": "cuit-01", "field": "计院 085404/085410 2026 数学科目", "unit": "计算机学院",
             "claims": [{"src": "N诺2026表/doubao深挖", "value": "数学一"}, {"src": "王道OCR 26 img10", "value": "数学二"}],
             "status": "待核", "action": "回看计院2026招生目录原文(jsjxy 3098公告附件)，9月中旬研招网2027目录一并复核"},
            {"id": "cuit-02", "field": "AI学院(019) 085410 2026 科目", "unit": "人工智能学院",
             "claims": [{"src": "官方2026招生目录(王道OCR img08)", "value": "英二+数一+408"}],
             "status": "已核实(2026)", "action": "2027 目录发布后复查是否改数二"},
            {"id": "cuit-03", "field": "2027 是否专硕改数二", "unit": "两学院085410",
             "claims": [{"src": "doubao 逐人数据册说明/同档横评", "value": "2027 专硕改数二，热度必反弹"},
                        {"src": "官方", "value": "未公告"}],
             "status": "待核", "action": "2026-09 中旬研招网 2027 硕士目录发布后核实"},
            {"id": "cuit-04", "field": "085411 分属两学院(软工院/应用数学院)", "unit": "全校",
             "claims": [{"src": "王道OCR/N诺", "value": "软工院 2026 一志愿73复71录；应用数学院一志愿2人均338(22408英二数二408)"}],
             "status": "已核实", "action": "报考/调剂写明学院"},
        ]
        # 各学院 2026 一志愿/调剂细节（人工核对结论，写进考情）
        obj["kaoqingDetail2026"] = obj.get("kaoqingDetail2026", []) + [
            {"college": "计算机学院", "program": "085410人工智能", "line2026": 264, "batch": "一志愿",
             "retestCnt": 2, "admitCnt": 1, "admitMin": 274, "admitMax": 276,
             "scope": "2人(274录取/276因综合面试0分未录)；逐人见专档xlsx", "url": "https://jsjxy.cuit.edu.cn/info/1062/3251.htm"},
            {"college": "计算机学院", "program": "085410人工智能", "line2026": 264, "batch": "调剂",
             "retestCnt": "25+33", "admitCnt": "18+7", "scope": "第一批18/25(266~323)；第二批7/33(315~341)"},
            {"college": "人工智能学院", "program": "085410人工智能", "line2026": 264, "batch": "一志愿",
             "retestCnt": 1, "admitCnt": 1, "admitMin": 271, "admitMax": 271,
             "scope": "唯一一志愿郭*(271)，录取；单科51/55/95/70", "url": "qkl.cuit.edu.cn 2765/2788号"},
            {"college": "人工智能学院", "program": "085410人工智能", "line2026": 264, "batch": "调剂",
             "retestCnt": "22+7", "admitCnt": "14+1", "scope": "第一批14/22(265~327)；第二批1/7(312)"},
        ]

    fn = sid(school) + ".json"
    p = os.path.join(SCH, fn)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
    meta_schools.append({"file": fn, "name": school, "nUnits": len(units_out),
                         "inCategory": units_out[0].get("category") if units_out else None,
                         "conflicts": len(obj["conflicts"])})

# 王道的完整链接（全部 487 校，按省归组后平铺；位置列解析）
wd_all = {}
wb = load_workbook(WD, read_only=True, data_only=True)
ws = wb["院校考情链接总表"]
for row in ws.iter_rows(values_only=True):
    vals = [norm(c) for c in row]
    if len(vals) >= 4 and vals[1] and vals[2]:
        prov, school = vals[1], vals[2]
        wd_all.setdefault(prov, {})[school] = {
            "type": vals[3],
            "count": vals[4] if len(vals) > 4 else None,
            "years": vals[5] if len(vals) > 5 else None,
            "links": wd_by_school.get(school, {}),
        }
wb.close()
os.makedirs(os.path.join(OUTDIR, "data"), exist_ok=True)
# N诺候补名单里不在主体/对照表的学校单独建档
known = set(schools.keys())
for cand in nn_cand:
    s = cand["school"]
    if s in known:
        continue
    if s not in {c["name"] for c in meta_schools}:
        obj = {
            "schema": "kaoyan-school/v1",
            "name": s,
            "units": [],
            "categoryGuess": "N诺新增候选(未入终极版S/C)",
            "nnCandidates": [c for c in nn_cand if c["school"] == s],
            "wangdaoLinks": wd_by_school.get(s, {}),
            "conflicts": [],
            "sources": collect_sources(s),
            "generated": {"by": "etl_build_db.py", "date": "2026-09-03",
                          "src": ["全国408_085410双非热度版_终极版_20260826.xlsx::N诺新增候选校",
                                  "王道考情链接总表_20260825.xlsx"],
                          "caution": "候选校：仅有N诺2026单条专业数据或王道链接，未做终极版主体档位评估。"},
        }
        fn = sid(s) + ".json"
        with open(os.path.join(SCH, fn), "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=1)
        meta_schools.append({"file": fn, "name": s, "nUnits": 0,
                             "inCategory": "NN候选", "conflicts": 0})

with open(os.path.join(OUTDIR, "data", "wangdao_links_all.json"), "w", encoding="utf-8") as f:
    json.dump({"note": "王道考研考情微信文章链接总表（2019-2026，机器整理）", "byProvince": wd_all},
              f, ensure_ascii=False, indent=1)

# 跨校冲突登记册（来源：06_终极版输出/核对报告_三源交叉_20260826.md + 本库校验）
REG = [
    {"school": "天津工业大学", "field": "2026 线口径", "claims": [{"src": "终极版", "value": "264(国家线)"}, {"src": "N诺复试线", "value": "332"}], "status": "待核", "action": "以官方复试细则为准"},
    {"school": "杭州电子科技大学", "field": "初试科目口径", "claims": [{"src": "终极版(22408范畴)", "value": "22408/英二数二"}, {"src": "N诺", "value": "英一数一408(11408)"}], "status": "警告-非22408", "action": "若目标 22408 请直接排除或改报其他方向"},
    {"school": "深圳大学", "field": "初试科目口径", "claims": [{"src": "终极版(22408范畴)", "value": "22408/英二数二"}, {"src": "N诺(人工智能学院/光明实验室/大湾区国际创新学院)", "value": "英一数一408(11408)"}], "status": "警告-非22408", "action": "若目标 22408 请直接排除或改报其他方向"},
    {"school": "山东科技大学", "field": "2026 线口径", "claims": [{"src": "终极版", "value": "264"}, {"src": "N诺", "value": "300"}], "status": "待核", "action": "以官方为准（可能学院/方向不同）"},
    {"school": "广东工业大学", "field": "2026 线口径", "claims": [{"src": "终极版", "value": "301"}, {"src": "N诺", "value": "337"}], "status": "待核", "action": "以官方为准（可能学院/方向不同）"},
    {"school": "河南大学", "field": "2026 线口径", "claims": [{"src": "终极版", "value": "264"}, {"src": "N诺", "value": "295"}], "status": "待核", "action": "以官方为准"},
    {"school": "北京信息科技大学", "field": "2026 线口径", "claims": [{"src": "终极版", "value": "264"}, {"src": "N诺", "value": "280"}], "status": "待核", "action": "以官方为准"},
    {"school": "广州大学", "field": "2026 线口径", "claims": [{"src": "终极版", "value": "310(AI院)/300(网安院)"}, {"src": "N诺", "value": "300"}], "status": "参考", "action": "学院/方向不同所致"},
    {"school": "四川师范大学", "field": "085404 录取人数", "claims": [{"src": "终极版v5", "value": "2"}, {"src": "川渝王道OCR", "value": "29"}], "status": "差异", "action": "口径批次不同，以官方拟录取名单为准"},
    {"school": "长春理工大学", "field": "N诺录取人数", "claims": [{"src": "N诺", "value": "322(疑似为录取均分类数值错位)"}], "status": "待核", "action": "复查原始数据"},
    {"school": "南京信息工程大学", "field": "2027 改考", "claims": [{"src": "2027改考动态(新东方2026-07)", "value": "408 → 817 信号与系统，反向改考"}], "status": "已官宣口径", "action": "勿按 408 复习该方向"},
    {"school": "成都信息工程大学", "field": "085410 科目/2027改数二", "claims": [{"src": "doubao深挖", "value": "2027 专硕改数二"}, {"src": "官方", "value": "未公告；2026 AI院085410=英二数一408"}], "status": "待核", "action": "9 月中旬研招网 2027 目录复核"},
]
os.makedirs(os.path.join(OUTDIR, "data"), exist_ok=True)
with open(os.path.join(OUTDIR, "data", "conflicts_registry.json"), "w", encoding="utf-8") as f:
    json.dump({"note": "跨校已知口径冲突/警告登记（v1；配合各校 school json 内 conflicts 使用）",
               "date": "2026-09-03", "items": REG}, f, ensure_ascii=False, indent=1)

with open(os.path.join(OUTDIR, "data", "meta.json"), "w", encoding="utf-8") as f:
    json.dump({"db": "kaoyan-db-ai-v1", "date": "2026-09-03", "nSchools": len(meta_schools),
               "schools": meta_schools}, f, ensure_ascii=False, indent=1)

print("schools:", len(meta_schools))
print("sample:", meta_schools[:5])
print("done")
