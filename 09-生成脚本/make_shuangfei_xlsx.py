# -*- coding: utf-8 -*-
"""从 全国408_085410双非热度版_20260820.html 提取数据，生成同结构 xlsx（7个Sheet）。"""
import io, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HTML = r"<SOURCE_DIR>\02_院校数据_原有\全国408_085410双非热度版_20260820.html"
OUTS = [
    r"<SOURCE_DIR>\02_院校数据_原有\全国408_085410双非热度版_20260820.xlsx",
    r"<MATERIAL_DIR>\02_院校数据_原有\全国408_085410双非热度版_20260820.xlsx",
]

s = io.open(HTML, encoding="utf-8").read()

def grab_array(begin_marker, end_marker):
    a = s.index(begin_marker) + len(begin_marker)
    b = s.index(end_marker, a) + 1
    return s[a:b]

S = json.loads(grab_array("var S=", "];\n// ===== 985/211"))
C = json.loads(grab_array("var C=", "];\n// ===== 初试科目映射"))
ka = s.index("var K={") + len("var K={") - 1
kb = s.index("};\nfunction exKey") + 1
K = json.loads(s[ka:kb])
O = json.loads(grab_array("var O=", "];\n// ===== 数据来源"))
SRCS = json.loads(grab_array("var SRCS=", "];\n\n// ===== 工具函数"))

def exKey(d):
    return K.get(d["n"] + "|" + d["c"]) or K.get(d["n"]) or "待确认"

ZONE_B = ["内蒙古", "广西", "海南", "贵州", "云南", "西藏", "甘肃", "青海", "宁夏", "新疆"]
def zoneOf(p): return "B" if p in ZONE_B else "A"
def baseLine(p): return 254 if zoneOf(p) == "B" else 264
def gapOf(d): return None if d["l"] is None else d["l"] - baseLine(d["p"])
def clamp(v): return max(0, min(100, v))
def lineScore(d):
    if d["l"] is None: return 45
    return clamp(round(gapOf(d) * 100 / 96))
def ratioScore(d):
    return 30 if d["ratio"] is None else clamp(round((d["ratio"] - 1) / 8 * 100))
def rrScore(d):
    return 35 if d["rr"] is None else clamp(round((d["rr"] - 1) / 1.5 * 100))
def fcScore(d):
    f = str(d["fc"])
    if any(x in f for x in ("大量调剂", "调剂", "一志愿不足", "缺口")): return 12
    if "一志愿为主" in f: return 72
    if f in ("待确认", "—"): return 45
    return 55
def compScore(d): return round(0.45 * lineScore(d) + 0.25 * ratioScore(d) + 0.2 * rrScore(d) + 0.1 * fcScore(d))
def heatScore(d): return round(0.7 * compScore(d) + 0.3 * (d["net"] or 0))
def gapStr(d):
    g = gapOf(d)
    return "—" if g is None else ("=国家线" if g == 0 else "+" + str(g))
def na(v): return "—" if v is None else v
def lh_str(d):
    a = list(d.get("lh") or [None, None, None]) + [d["l"]]
    return " / ".join("—" if v is None else str(v) for v in a)

wb = openpyxl.Workbook()
HDR_FILL = PatternFill("solid", fgColor="0F766E")
HDR_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="C8D6D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

def sheet(ws, headers, rows, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY; cell.border = BORDER; cell.alignment = WRAP
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# 1 总览
rows = []
for d in S:
    rows.append([d["n"], d["p"], d["r"], d["t"], d["c"], exKey(d), d["d"], d["l"], lh_str(d), gapStr(d), d["plan"],
                 d["fc"], na(d["ratio"]), na(d["rr"]), na(d.get("rec")), na(d.get("adm")), na(d.get("max_s")),
                 na(d.get("min_s")), na(d.get("avg_s")), d.get("scope") or "—", d["ai"], d["src"], d["net"], heatScore(d), d["note"]])
ws = wb.active; ws.title = "总览"
sheet(ws, ["院校","省份","大区","层次","学院","初试科目","方向","2026复试线","历年线(23/24/25/26)","线差","拟招","一志愿/调剂","报录比","复录比","复试人数","录取人数","录取最高","录取最低","录取平均","口径","AI方向备注","来源","网络热度","综合热度","备注"], rows,
      [16,6,8,12,30,10,22,10,24,8,14,16,8,8,9,9,9,9,9,46,16,12,8,8,50])

# 2 热度榜
sortedS = sorted(S, key=heatScore, reverse=True)
rows = [[i+1, d["n"], d["r"], d["t"], d["c"], exKey(d), d["l"], gapStr(d), compScore(d), d["net"], heatScore(d), d["ai"], d["src"], d["note"]]
        for i, d in enumerate(sortedS)]
ws = wb.create_sheet("热度榜")
sheet(ws, ["排名","院校","大区","层次","学院","初试科目","2026线","线差","竞争热度","网络热度","综合热度","AI方向备注","来源","备注"], rows,
      [6,16,8,12,30,10,9,9,9,9,9,16,12,50])

# 3 区域榜
E_REG = ["华北","华东","东北"]; W_REG = ["华中","华南","西南","西北"]
rows = []
for label, regs in (("华东/华北/东北", E_REG), ("华中/华南/西南/西北", W_REG)):
    merged = [("主体", d) for d in S if d["r"] in regs] + [("对照", d) for d in C if d["r"] in regs]
    merged.sort(key=lambda x: heatScore(x[1]), reverse=True)
    for k, (cat, d) in enumerate(merged, 1):
        rows.append([label, cat, k, d["n"], d["p"], d["r"], d["t"], d["c"], exKey(d), d["l"], gapStr(d),
                     compScore(d), d["net"], heatScore(d), d["ai"], d["src"], d["note"]])
ws = wb.create_sheet("区域榜")
sheet(ws, ["榜单","类别","排名","院校","省份","大区","层次","学院","初试科目","2026线","线差","竞争热度","网络热度","综合热度","AI方向备注","来源","备注"], rows,
      [20,6,6,16,6,8,12,30,10,9,9,9,9,9,16,12,50])

# 4 招录比对比
rows = [[d["n"], d["r"], d["t"], d["c"], exKey(d), d["l"], gapStr(d), d["plan"], d["fc"], na(d["ratio"]), na(d["rr"]),
         na(d.get("rec")), na(d.get("adm")), na(d.get("max_s")), na(d.get("min_s")), na(d.get("avg_s")), compScore(d), d["src"]]
        for d in sortedS]
ws = wb.create_sheet("招录比对比")
sheet(ws, ["院校","大区","层次","学院","初试科目","2026线","线差","拟招","一志愿/调剂","报录比","复录比","复试人数","录取人数","录取最高","录取最低","录取平均","竞争热度","来源"], rows,
      [16,8,12,30,10,9,9,14,16,8,8,9,9,9,9,9,9,12])

# 4.5 考情明细
rows = []
for d in S + C:
    if any(d.get(k) is not None for k in ("rec", "adm", "max_s", "min_s", "avg_s")) or d.get("scope"):
        rows.append([d["n"], d["c"], d["t"], exKey(d), na(d["l"]), na(d["plan"]), na(d.get("rec")), na(d.get("adm")),
                     na(d.get("rr")), na(d.get("ratio")), na(d.get("max_s")), na(d.get("min_s")), na(d.get("avg_s")),
                     d.get("scope") or "—", d.get("verified") or "查无", d.get("srcu") or "—"])
ws = wb.create_sheet("考情明细")
sheet(ws, ["院校","学院","层次","初试科目","2026线","拟招","复试人数","录取人数","复录比","报录比","录取最高","录取最低","录取平均","口径","核实状态","来源URL"], rows,
      [18,30,12,12,9,14,9,9,8,8,9,9,9,60,12,70])

# 5 调剂院校
adj = [d for d in sortedS if any(x in str(d["fc"]) for x in ("调剂", "一志愿不足", "缺口"))]
rows = [[d["n"], exKey(d), d["r"], d["t"], d["l"], d["fc"], d["ai"], d["src"]] for d in adj]
ws = wb.create_sheet("调剂院校")
sheet(ws, ["院校","初试科目","大区","层次","2026线","一志愿/调剂","AI方向","来源"], rows,
      [16,10,8,12,9,18,16,12])

# 6 重点院校（Top 12，按综合热度，同校多学院合并取最高）
bySchool = {}
for d in S:
    bySchool.setdefault(d["n"], []).append(d)
top = sorted(bySchool.items(), key=lambda kv: max(heatScore(x) for x in kv[1]), reverse=True)[:12]
rows = []
for name, ds in top:
    main = max(ds, key=heatScore)
    rows.append([name, main["p"], main["r"], main["t"], main["c"], exKey(main), main["d"], main["l"], main["plan"],
                 main["fc"], main["ai"], main["src"], heatScore(main), "；".join(d["note"] for d in ds)[:200]])
ws = wb.create_sheet("重点院校")
sheet(ws, ["院校","省份","大区","层次","学院","初试科目","方向","2026线","拟招","一志愿/调剂","AI方向","来源","综合热度","备注"], rows,
      [16,6,8,12,30,10,22,9,14,16,16,12,9,50])

# 7 数据源
rows = [[t, u] for t, u in SRCS]
ws = wb.create_sheet("数据源")
sheet(ws, ["来源","URL"], rows, [42, 90])

for out in OUTS:
    wb.save(out)
    print("已写入:", out)

print("S:", len(S), "C:", len(C), "O:", len(O), "SRCS:", len(SRCS), "调剂:", len(adj))
