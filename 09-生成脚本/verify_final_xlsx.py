# -*- coding: utf-8 -*-
"""终极版 xlsx 完整性校验"""
import openpyxl

P = r'<SOURCE_DIR>\02_院校数据_原有\全国408_085410双非热度版_终极版_20260824.xlsx'
wb = openpyxl.load_workbook(P, read_only=True)
print('Sheet 列表:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  {name}: {ws.max_row} 行 x {ws.max_column} 列')
