# -*- coding: utf-8 -*-
"""
三源整合进终极版 HTML（2026-08-26）
数据源：1) 王道考情链接总表(487校/2477链接) 2) N诺双非408考情(108校/469行) 3) 川渝王道OCR(30校/184行)
输出：06_终极版输出/全国408_085410双非热度版_终极版_20260826.html（20260824 原版不动）
"""
import io, json, re, collections
import openpyxl

SRC_HTML = r"06_终极版输出\全国408_085410双非热度版_终极版_20260824.html"
OUT_HTML = r"06_终极版输出\全国408_085410双非热度版_终极版_20260826.html"

def norm(s):
    if s is None: return ""
    return re.sub(r"\s+", "", str(s).strip())

def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def num(v):
    if not is_num(v): return None
    x = float(v)
    return int(x) if x == int(x) else round(x, 1)

def rows_of(ws):
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]

# ---------------- 加载三源 ----------------
wd = openpyxl.load_workbook(r"deliverables\20260825-王道考情链接总表\王道考情链接总表_20260825.xlsx", read_only=True, data_only=True)
wd_links = collections.defaultdict(list)
for r in rows_of(wd["链接明细"])[1:]:
    if len(r) < 6 or not r[2]: continue
    wd_links[norm(r[2])].append((norm(r[4]), norm(r[5])))
wd_sum = {}
for r in rows_of(wd["院校考情链接总表"])[1:]:
    if len(r) < 7 or not r[2]: continue
    wd_sum[norm(r[2])] = {"cnt": r[4], "years": norm(r[5])}
wd.close()

cy = openpyxl.load_workbook(r"deliverables\20260825-川渝王道考情提取\川渝王道考情提取_20260825.xlsx", read_only=True, data_only=True)
cy_rows = collections.defaultdict(list)
hdr = None
for r in rows_of(cy["川渝王道考情提取"]):
    if hdr is None: hdr = r; continue
    if len(r) < 11 or not r[0]: continue
    cy_rows[norm(r[0])].append(r)
cy.close()

nn = openpyxl.load_workbook(r"deliverables\20260826-双非408考情_N诺提取\双非408考情_2026提取.xlsx", read_only=True, data_only=True)
nn_rows = collections.defaultdict(list)
hdr = None
for r in rows_of(nn["2026年408系明细"]):
    if hdr is None: hdr = r; continue
    if len(r) < 19 or not r[0]: continue
    nn_rows[norm(r[0])].append(r)
nn.close()

# ---------------- 选取逻辑 ----------------
def pick_nn(rows):
    def score(r):
        maj, subj = norm(r[3]), norm(r[4])
        s = 0
        if "人工智能" in maj: s += 300
        elif "智能科学" in maj or "智能技术" in maj: s += 250
        elif "电子信息" in maj: s += 200
        elif "计算机" in maj: s += 150
        elif "软件" in maj: s += 100
        if subj in ("英二数二408", "推测408"): s += 80
        elif subj == "英一数一408": s -= 40
        if is_num(r[9]): s += 20
        if is_num(r[5]): s += 10
        if is_num(r[6]): s += min(r[6], 800) / 100.0
        return s
    return max(rows, key=score)

def nn_dict(r):
    return {
        "dept": norm(r[2]), "maj": norm(r[3]), "subj": norm(r[4]),
        "line": num(r[5]), "rc": num(r[6]), "adj": num(r[7]),
        "tot": num(r[8]), "a408": num(r[9]),
        "pol": num(r[10]), "eng": num(r[11]), "math": num(r[12]),
        "ad": num(r[13]), "adavg": num(r[14]),
        "lr": num(r[15]), "br": num(r[17]),
    }

def pick_cy(rows26):
    def has(r): return any(r[i] not in (None, "") for i in range(3, 9))
    withdata = [r for r in rows26 if has(r)]
    if not withdata: return None
    order = {"085410": 0, "085404": 1, "085400": 2, "085411": 3, "081200": 4}
    return min(withdata, key=lambda r: (order.get(norm(r[2]), 9), norm(r[2])))

def cy_dict(r):
    return {
        "code": norm(r[2]), "line": num(r[3]), "rc": num(r[4]), "ad": num(r[5]),
        "mx": num(r[6]), "mn": num(r[7]), "avg": num(r[8]),
        "seg": norm(r[9]), "note": norm(r[10]),
    }

def wd_dict(name):
    links = wd_links.get(name)
    if not links: return None
    years = sorted(set(y for y, u in links), key=lambda y: int(re.sub(r"\D", "", y)))
    u26 = next((u for y, u in links if y == "26考情"), None)
    if u26 is not None:
        y = "26考情"
    else:
        y, u26 = links[-1]
    return {"cnt": len(links), "years": ",".join(years), "y": y, "u": u26}

# ---------------- 解析原 HTML ----------------
s = io.open(SRC_HTML, encoding="utf-8").read()

def grab(begin, end):
    a = s.index(begin) + len(begin)
    b = s.index(end, a) + 1
    return s[a:b]

S = json.loads(grab("var S=", "];\n// ===== 985/211"))
C = json.loads(grab("var C=", "];\n// ===== 初试科目映射"))

# ---------------- 富化 S / C ----------------
for d in S + C:
    n = norm(d["n"])
    rows = nn_rows.get(n)
    if rows:
        d["nn"] = nn_dict(pick_nn(rows))
    wd = wd_dict(n)
    if wd:
        d["wd"] = wd
    c26 = [r for r in cy_rows.get(n, []) if norm(r[1]) == "26考情"]
    pc = pick_cy(c26)
    if pc is not None:
        d["cy"] = cy_dict(pc)

def fmt_block(arr):
    return "var S=[\n" + ",\n".join(json.dumps(x, ensure_ascii=False) for x in arr) + "\n];"

# 替换 S 块
a = s.index("var S=[")
b = s.index("];\n// ===== 985/211", a) + 2
s = s[:a] + "var S=[" + ",\n".join(json.dumps(x, ensure_ascii=False) for x in S) + "\n];" + s[b:]
# 替换 C 块
a = s.index("var C=[")
b = s.index("];\n// ===== 初试科目映射", a) + 2
s = s[:a] + "var C=[" + ",\n".join(json.dumps(x, ensure_ascii=False) for x in C) + "\n];" + s[b:]

# ---------------- OV_COLS 增列 ----------------
a_ov = s.index("var OV_COLS=[")
b_ov = s.index("];\nvar ov={", a_ov) + 2
old_ov = s[a_ov:b_ov]
new_ov = (
    "var OV_COLS=[\n"
    '  ["院校","#",0],["省份","#",1],["大区","#",2],["层次","#",3],["学院","#",4],\n'
    '  ["初试科目","#",5],["方向","#",6],["2026复试线","num",7],["历年线(23/24/25/26)","#",8],["线差","num",9],["拟招","#",10],\n'
    '  ["一志愿/调剂","#",11],["报录比","num",12],["复录比","num",13],\n'
    '  ["复试人数","num",14],["录取人数","num",15],["录取最高","num",16],["录取最低","num",17],["录取平均","num",18],["口径","#",19],\n'
    '  ["AI方向备注","#",20],["来源","#",21],["网络热度","num",22],\n'
    '  ["408均分(N诺)","num",24],["录取率(N诺)","num",25],["王道考情","#",26],\n'
    '  ["备注","#",23]\n'
    "];"
)
assert s.count(old_ov) == 1, "OV_COLS 未找到"
s = s.replace(old_ov, new_ov, 1)

# ---------------- 新 JS 帮助函数（插在 全国总览表 之前） ----------------
helpers = '''
// ===== 三源整合（2026-08-26 新增）=====
function nnInfo(d){return d.nn||null;}
function nn408Cell(d){
  var x=nnInfo(d);
  if(!x||x.a408==null)return '<td class="num"><span class="na">—</span></td>';
  var warn=(x.subj!=="英二数二408"&&x.subj!=="推测408")?' <span class="tag-11408" title="N诺口径：'+esc(x.subj)+'，非22408">'+esc(x.subj)+'</span>':'';
  return '<td class="num nn-hl" title="N诺2026 · '+esc(x.dept)+' · '+esc(x.maj)+' · '+esc(x.subj)+'">'+x.a408+warn+"</td>";
}
function lrCell(d){
  var x=nnInfo(d);
  if(!x||x.lr==null)return '<td class="num"><span class="na">—</span></td>';
  return '<td class="num" title="N诺录取率=录取/复试">'+Math.round(x.lr*100)+"%</td>";
}
function wdCell(d){
  if(!d.wd||!d.wd.u)return '<td><span class="na">—</span></td>';
  return '<td><a class="wd-link" href="'+esc(d.wd.u)+'" target="_blank" rel="noopener" title="王道考情共'+d.wd.cnt+'条（覆盖 '+(d.wd.years||"")+'）">'+esc(d.wd.y)+' ↗</a></td>';
}
'''
s = s.replace("// ===== 全国总览表 =====", helpers + "// ===== 全国总览表 =====", 1)

# ---------------- ovRow 尾部加 3 列 ----------------
old_tail = '''    '<td class="num">'+d.net+"</td>"+
    '<td class="score-cell">'+esc(d.note)+"</td></tr>";
}function ovPass(i){'''
new_tail = '''    '<td class="num">'+d.net+"</td>"+
    nn408Cell(d)+
    lrCell(d)+
    wdCell(d)+
    '<td class="score-cell">'+esc(d.note)+"</td></tr>";
}function ovPass(i){'''
assert s.count(old_tail) == 1
s = s.replace(old_tail, new_tail, 1)

# ---------------- ovSortKey 加 24/25 ----------------
old_sk = '    if(ov.sort===22)return d.net==null?-9999:d.net;'
new_sk = old_sk + '\n    if(ov.sort===24)return d.nn&&d.nn.a408!=null?d.nn.a408:-9999;\n    if(ov.sort===25)return d.nn&&d.nn.lr!=null?d.nn.lr:-9999;'
assert s.count(old_sk) == 1
s = s.replace(old_sk, new_sk, 1)

# ---------------- CSS ----------------
old_css = "</style>"
new_css = """
.nn-hl{background:#eff6ff!important;color:#1d4ed8;font-weight:700}
.tag-11408{font-size:10px;color:#b91c1c;border:1px solid #fecaca;border-radius:4px;padding:0 3px;background:#fef2f2;font-weight:600}
.wd-link{color:var(--accent);text-decoration:none;border-bottom:1px dashed var(--accent);white-space:nowrap}
.three-src{border-left:6px solid #2563eb;padding-left:12px}
.three-src table{min-width:1180px}
.three-src td{font-size:12px}
.three-src .zj-ok{color:var(--ok);font-weight:700}
.three-src .zj-warn{color:var(--bad);font-weight:600}
.three-src .zj-info{color:var(--muted)}
</style>"""
assert s.count(old_css) == 1
s = s.replace(old_css, new_css, 1)

# ---------------- 三源核对区块（插在 recommender 前） ----------------
def cy_cell(d):
    c = d.get("cy")
    if not c: return '<span class="zj-info">—</span>'
    parts = []
    if c.get("rc") is not None: parts.append("复试" + str(c["rc"]))
    if c.get("ad") is not None: parts.append("录" + str(c["ad"]))
    if c.get("avg") is not None: parts.append("均" + str(c["avg"]))
    body = c["code"] + ((" " + "·".join(parts)) if parts else " 无数据")
    return '<span title="川渝王道OCR 2026考情，专业口径：' + esc_html(c["code"]) + '">' + body + "</span>"

def flag_cell(d):
    flags = []
    nn_ = d.get("nn")
    if nn_:
        if nn_["subj"] not in ("英二数二408", "推测408"):
            flags.append('⚠️N诺口径' + esc_html(nn_["subj"]))
        if nn_.get("line") and d.get("l") and abs(nn_["line"] - d["l"]) >= 5:
            diff = nn_["line"] - d["l"]
            flags.append("⚠️N诺线差" + ("+" if diff > 0 else "") + str(diff))
    cy_ = d.get("cy")
    if cy_:
        if cy_["code"] != "085410":
            flags.append("参考" + cy_["code"])
        if d.get("rec") is not None and cy_.get("rc") is not None and d["rec"] != cy_["rc"]:
            flags.append("⚠️复试人数差异(终极" + str(d["rec"]) + "/王道" + str(cy_["rc"]) + ")")
        if d.get("adm") is not None and cy_.get("ad") is not None and d["adm"] != cy_["ad"]:
            flags.append("⚠️录取人数差异")
    return "；".join(flags) if flags else '<span class="zj-info">—</span>'

def esc_html(x):
    return (str(x).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))

rows_html = []
for tag, lst in (("主体", S), ("对照", C)):
    for d in lst:
        nn_ = d.get("nn")
        nn_cell = "—"
        if nn_:
            subj = "" if nn_["subj"] in ("英二数二408", "推测408") else ' <span class="tag-11408">' + esc_html(nn_["subj"]) + "</span>"
            lr = "" if nn_.get("lr") is None else str(round(nn_["lr"] * 100)) + "%"
            nn_cell = esc_html(nn_["maj"]) + "<br><span class='zj-info'>" + esc_html(nn_["dept"]) + " · " + esc_html(nn_["subj"]) + "</span>"
            a408 = str(nn_["a408"]) if nn_.get("a408") is not None else "—"
        else:
            a408 = lr = "—"
        wd_ = d.get("wd")
        wd_cell = "—" if not wd_ else '<a class="wd-link" href="' + esc_html(wd_["u"]) + '" target="_blank" rel="noopener">' + esc_html(wd_["y"]) + "（共" + str(wd_["cnt"]) + "条）</a>"
        rows_html.append(
            "<tr><td><b>" + esc_html(d["n"]) + "</b></td><td>" + tag + "</td><td>" + esc_html(d["t"]) + "</td>"
            "<td class='num'>" + esc_html(d.get("l")) + "</td>"
            "<td>" + (esc_html(nn_["maj"]) if nn_ else "—") + "<br><span class='zj-info'>" + (esc_html(nn_["dept"]) if nn_ else "") + "</span></td>"
            "<td class='num nn-hl'>" + a408 + (subj if nn_ else "") + "</td>"
            "<td class='num'>" + lr + "</td>"
            "<td>" + cy_cell(d) + "</td>"
            "<td>" + wd_cell + "</td>"
            "<td>" + flag_cell(d) + "</td></tr>"
        )

section = """
<section id="three-src" class="three-src">
  <h2>五·五、三源交叉核对表（新增 2026-08-26）<span class="hint">王道考情链接 × N诺408考情 × 川渝王道OCR · 与终极版逐校对齐</span></h2>
  <div class="banner tip" style="margin:10px 0">
    <b>三源整合说明：</b>① <b>王道考情链接</b>（腾讯文档 2019-2026，487 校 2477 条微信文章链接）逐校挂到总览表「王道考情」列，点链接直达当年考情分析；② <b>N诺双非408考情</b>（108 校 469 行 2026 明细）补「408均分/政治/英语/数学均分/录取率」维度，录取率=录取/复试，其「报录比」实为复试/录取口径（即复录比）；③ <b>川渝王道OCR</b>（30 校 184 行，2025/2026 两届，复试线识别率有限）补川渝校复试/录取人数与均分。<span class="zj-warn">⚠️ 标注「11408」=N诺该行初试为英一数一408，与你的 22408 口径不符，仅作参考；「推测408」=N诺未确认科目。</span>
  </div>
  <div class="table-wrap">
    <table class="three-src">
      <thead><tr><th>院校</th><th>类别</th><th>层次</th><th>终极版2026线</th><th>N诺专业（院系）</th><th>N诺408均分</th><th>N诺录取率</th><th>川渝王道26考情</th><th>王道考情链接</th><th>标注</th></tr></thead>
      <tbody>
""" + "\n".join(rows_html) + """
      </tbody>
    </table>
  </div>
</section>

"""
marker = '<section id="recommender">'
assert s.count(marker) == 1
s = s.replace(marker, section + marker, 1)

# ---------------- SRCS 追加 3 条 ----------------
old_src = '["海南大学2026复试细则", "https://sice.hainanu.edu.cn/info/1026/9883.htm"]\n];'
new_src = (old_src.replace("\n];", ",\n") +
    '["王道考情2019-2026汇总(腾讯文档·487校2477条)","https://docs.qq.com/smartsheet/DTFNGVXd5aE9KZVRL?tab=MIImLX&viewId=v3rKFu"],\n' +
    '["N诺考研(noobdream)双非院校408考情2026","https://www.noobdream.com/"],\n' +
    '["川渝王道考情OCR提取(2026-08-25,30校184行)","deliverables/20260825-川渝王道考情提取/"]\n];')
assert s.count(old_src) == 1
s = s.replace(old_src, new_src, 1)

# ---------------- 标题 / 日期 / banner / footer ----------------
s = s.replace("（2026-08-24汇总）", "（2026-08-26 三源整合版）", 1)
s = s.replace("<b>2026-08-21</b><span>数据截止</span>", "<b>2026-08-26</b><span>数据截止</span>", 1)
s = s.replace("截至 2026-08-21，研招网 2027 硕士专业目录尚未发布，9 月发布后请逐校复核）。",
              "截至 2026-08-26，研招网 2027 硕士专业目录尚未发布，9 月发布后请逐校复核）。2026-08-26 已整合三源新数据（王道考情链接全表 / N诺双非408考情 / 川渝王道OCR），详见下方「五·五」节。", 1)
s = s.replace("未编造），报考请以官方 2027 招生简章为准。",
              "未编造；2026-08-26 三源整合：新增王道考情链接、N诺408均分/录取率、川渝王道OCR 交叉核对，详见「五·五」节），报考请以官方 2027 招生简章为准。", 1)

io.open(OUT_HTML, "w", encoding="utf-8").write(s)
print("已写出:", OUT_HTML)
print("S:", len(S), "C:", len(C))
print("S 命中: wd=%d nn=%d cy=%d" % (
    sum(1 for d in S if d.get("wd")), sum(1 for d in S if d.get("nn")), sum(1 for d in S if d.get("cy"))))
