# -*- coding: utf-8 -*-
"""
408 考研终极版生成脚本
综合三份资料：
  1. trae 川渝 408 院校对比（E:\\some work\\川渝408.xlsx）
  2. 我的全国双非 085410 数据（全国408_双非院校_完整版_20260813.xlsx）
  3. 豆包 408 全套资料（Downloads 的 .md）
产出：终极 HTML（带导航）+ 终极 xlsx
"""
import glob
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

from md_conv import md_to_html, esc


def read_sheet_rows(path, sheet_index):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if c is None else str(c)) for c in row])
    return ws.title, rows


def rows_to_html(rows, skip_empty=True):
    """把二维列表转成 HTML 表格。"""
    if not rows:
        return ""
    header = rows[0]
    body = rows[1:]
    html = '<div class="table-scroll"><table><thead><tr>'
    html += "".join(f"<th>{esc(c)}</th>" for c in header)
    html += "</tr></thead><tbody>"
    for r in body:
        if skip_empty and not any(str(c).strip() for c in r):
            continue
        html += "<tr>" + "".join(f"<td>{esc(c)}</td>" for c in r) + "</tr>"
    html += "</tbody></table></div>"
    return html


# ---------------------------------------------------------------------------
# 一、读取豆包 md，按篇分割
# ---------------------------------------------------------------------------
DOUBAO = r"<DOWNLOADS_DIR>\408考研全套资料·院校大全+复习指南+核心笔记.md"
with open(DOUBAO, encoding="utf-8") as f:
    db_text = f.read()

# 按 # 第X篇 分割
part_marks = re.finditer(r"^# (第[一二三四五]篇 .+)$", db_text, flags=re.M)
parts = []
prev_end = None
part_titles = []
for m in part_marks:
    if prev_end is not None:
        parts.append((part_titles[-1], db_text[prev_end:m.start()]))
    part_titles.append(m.group(1))
    prev_end = m.start()
if prev_end is not None:
    parts.append((part_titles[-1], db_text[prev_end:]))

print("豆包分篇数:", len(parts))
for title, _ in parts:
    print("  -", title)

# ---------------------------------------------------------------------------
# 二、读取 trae 川渝 xlsx
# ---------------------------------------------------------------------------
TRAE = glob.glob(r"E:\some work\川渝408.xlsx")[0]
trae_wb = openpyxl.load_workbook(TRAE, data_only=True)
trae_sheets = []
for ws in trae_wb.worksheets:
    rows = [["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    trae_sheets.append((ws.title, rows))
print("trae sheets:", [t for t, _ in trae_sheets])

# ---------------------------------------------------------------------------
# 三、读取我的双非 xlsx
# ---------------------------------------------------------------------------
MINE = glob.glob("*双非院校_完整版*.xlsx")[0]
mine_wb = openpyxl.load_workbook(MINE, data_only=True)


def mine_rows(sheet_name):
    ws = mine_wb[sheet_name]
    return [["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)]


print("mine sheets:", mine_wb.sheetnames)


# ---------------------------------------------------------------------------
# 四、为我的双非数据生成带热力色阶的 HTML
# ---------------------------------------------------------------------------
def heat_cell(val, low=20, high=70):
    try:
        v = float(val)
    except (TypeError, ValueError):
        return f"<td>{esc(val)}</td>"
    t = max(0.0, min(1.0, (v - low) / (high - low)))
    if t < 0.33:
        color = "rgb(235,245,255)"
    elif t < 0.66:
        color = "rgb(255,242,204)"
    elif t < 0.85:
        color = "rgb(255,214,153)"
    else:
        color = "rgb(255,170,150)"
    return f'<td style="background:{color};font-weight:600">{esc(val)}</td>'


def line_cell(val):
    try:
        v = int(val)
    except (TypeError, ValueError):
        return f"<td>{esc(val)}</td>"
    if v >= 320:
        color = "#c0392b"
    elif v >= 280:
        color = "#e67e22"
    else:
        color = "#1e8449"
    return f'<td style="color:{color};font-weight:700">{esc(val)}</td>'


def rel_tag(val):
    if "已核实" in str(val):
        return '<span class="tag rel-ok">已核实</span>'
    return '<span class="tag rel-ref">参考</span>'


def mine_overview_html():
    rows = mine_rows("总览")
    if not rows:
        return ""
    header = rows[0]
    # 列索引
    idx = {h: i for i, h in enumerate(header)}
    # 综合热度、复试线、可靠度、可报专业
    html = '<div class="table-scroll"><table><thead><tr>'
    html += "".join(f"<th>{esc(c)}</th>" for c in header)
    html += "</tr></thead><tbody>"
    for r in rows[1:]:
        if not any(str(c).strip() for c in r):
            continue
        html += "<tr>"
        for i, c in enumerate(r):
            col = header[i]
            if col == "综合热度":
                html += heat_cell(c)
            elif col == "2026复试线":
                html += line_cell(c)
            elif col == "专业可靠度":
                html += f"<td>{rel_tag(c)}</td>"
            elif col == "院校":
                html += f'<td class="school">{esc(c)}</td>'
            elif col in ("可报专业(408)", "备注"):
                html += f'<td class="majors">{esc(c)}</td>'
            else:
                html += f"<td>{esc(c)}</td>"
        html += "</tr>"
    html += "</tbody></table></div>"
    return html


def mine_majors_html():
    rows = mine_rows("408专业速查")
    if not rows:
        return ""
    header = rows[0]
    html = '<div class="table-scroll"><table><thead><tr>'
    html += "".join(f"<th>{esc(c)}</th>" for c in header)
    html += "</tr></thead><tbody>"
    for r in rows[1:]:
        if not any(str(c).strip() for c in r):
            continue
        html += "<tr>"
        for i, c in enumerate(r):
            if header[i] == "专业代码":
                html += f'<td class="code">{esc(c)}</td>'
            elif header[i] == "学位类型":
                cls = "deg-xue" if "学硕" in str(c) else "deg-zhuan"
                html += f'<td><span class="tag {cls}">{esc(c)}</span></td>'
            else:
                html += f"<td>{esc(c)}</td>"
        html += "</tr>"
    html += "</tbody></table></div>"
    return html


def mine_adjust_html():
    rows = mine_rows("调剂院校")
    return rows_to_html(rows)


# ---------------------------------------------------------------------------
# 五、组装终极 HTML
# ---------------------------------------------------------------------------
CSS = (
    "<style>"
    ":root{--blue:#1F4E78;--blue2:#2E75B6;--bg:#f4f6f9;--card:#fff;--line:#e3e8ef;--text:#1f2933;--muted:#6b7280}"
    "*{box-sizing:border-box}html{scroll-behavior:smooth}"
    "body{margin:0;font-family:-apple-system,'Segoe UI','PingFang SC','Microsoft YaHei',sans-serif;background:var(--bg);color:var(--text);line-height:1.7}"
    ".wrap{max-width:1180px;margin:0 auto;padding:16px}"
    "header{background:linear-gradient(135deg,#1F4E78,#2E75B6);color:#fff;padding:30px 18px}"
    "header h1{margin:0 0 6px;font-size:26px}header p{margin:0;opacity:.92;font-size:14px}"
    ".pill{display:inline-block;background:rgba(255,255,255,.18);border-radius:14px;padding:2px 10px;font-size:12px;margin:8px 6px 0 0}"
    "nav.toc{position:sticky;top:0;z-index:10;background:#fff;border-bottom:1px solid var(--line);padding:8px 0;overflow-x:auto}"
    "nav.toc .wrap{display:flex;gap:8px;flex-wrap:nowrap}"
    "nav.toc a{flex:0 0 auto;text-decoration:none;color:var(--blue);font-size:13px;padding:5px 12px;border-radius:16px;background:#eef4fb;white-space:nowrap}"
    "nav.toc a:hover{background:var(--blue);color:#fff}"
    "section{margin-top:26px}h1,h2,h3,h4{color:var(--blue);line-height:1.4}"
    "h2{font-size:21px;border-bottom:2px solid var(--blue2);padding-bottom:6px;margin:28px 0 12px}"
    "h3{font-size:17px;margin:20px 0 8px}h4{font-size:15px;margin:16px 0 6px}p{margin:8px 0}"
    ".note-block{background:#fff8e6;border:1px solid #f0d9a0;border-radius:10px;padding:12px 14px;font-size:14px;color:#7a5b13}"
    ".legend{display:flex;flex-wrap:wrap;gap:8px;margin:8px 0;font-size:13px;color:var(--muted)}"
    ".legend span{display:inline-flex;align-items:center;gap:5px}"
    ".dot{width:12px;height:12px;border-radius:3px;display:inline-block}"
    "table{width:100%;border-collapse:collapse;background:var(--card);font-size:13px;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06)}"
    "th{background:var(--blue);color:#fff;padding:9px 7px;text-align:center;font-weight:600;white-space:nowrap;position:sticky;top:0}"
    "td{padding:8px 7px;border-bottom:1px solid var(--line);text-align:center;vertical-align:top}"
    "tr:nth-child(even) td{background:#f8fafc}tbody tr:hover td{background:#eef4fb}"
    ".school{font-weight:600;text-align:left;white-space:nowrap}"
    ".majors{font-size:12px;color:var(--muted);text-align:left;white-space:normal}"
    ".note{text-align:left;color:var(--muted);font-size:12px}"
    ".code{font-family:ui-monospace,Consolas,monospace;font-weight:700;color:var(--blue2)}"
    ".tag{display:inline-block;border-radius:10px;padding:1px 8px;font-size:11px;white-space:nowrap}"
    ".deg-xue{background:#e8f0fe;color:#1a56a8}.deg-zhuan{background:#e6f7ee;color:#16794a}"
    ".rel-ok{background:#e6f7ee;color:#16794a}.rel-ref{background:#fef3e2;color:#a8641a}"
    ".rank{font-weight:700;color:var(--blue2)}"
    ".table-scroll{overflow-x:auto;border-radius:10px;margin:8px 0}"
    "blockquote{border-left:4px solid var(--blue2);background:#eef4fb;margin:10px 0;padding:8px 14px;color:#41546b;border-radius:0 8px 8px 0}"
    "pre{background:#1f2933;color:#e5e7eb;padding:14px;border-radius:10px;overflow-x:auto;font-size:13px;line-height:1.5}"
    "code{background:#eef1f5;color:#b91c1c;padding:1px 5px;border-radius:4px;font-family:ui-monospace,Consolas,monospace;font-size:.9em}"
    "pre code{background:none;color:inherit;padding:0}ul,ol{padding-left:24px}li{margin:4px 0}"
    "footer{text-align:center;color:var(--muted);font-size:12px;margin:32px 0 12px}"
    "@media(max-width:720px){header h1{font-size:20px}th,td{font-size:12px;padding:6px 4px}.wrap{padding:10px}}"
    "</style>"
)

# 豆包各篇转 HTML
db_parts_html = [(title, md_to_html(body)) for title, body in parts]

nav_items = []
for i, (title, _) in enumerate(db_parts_html, start=1):
    nav_items.append(f'<a href="#part{i}">{esc(title)}</a>')

part1_body = db_parts_html[0][1] if db_parts_html else ""

trae_html_blocks = []
for title, rows in trae_sheets:
    trae_html_blocks.append(f"<h3>{esc(title)}</h3>\n{rows_to_html(rows)}")
trae_all = "\n".join(trae_html_blocks)

mine_all = (
    "<h3>全国双非院校 085410 总览（专业级，含可报专业与可靠度）</h3>"
    + mine_overview_html()
    + "<h3>调剂友好院校</h3>"
    + mine_adjust_html()
)

part1_content = (
    "<h2>408 专业速查（已核实）</h2>"
    + mine_majors_html()
    + part1_body
    + "<h2>川渝院校专业级数据（trae 深度版）</h2>"
    + "<p class='note-block'>川渝地区院校的专业级数据，覆盖总览、招录比、难度评级、决策矩阵、就业成本、复试对比、调剂、分数线趋势、报考建议九个维度，信息粒度比概览名单更细。</p>"
    + trae_all
    + "<h2>全国双非院校 085410 精细数据（帅帅整理）</h2>"
    + "<p class='note-block'>聚焦「考 408 的双非院校」的人工智能专硕(085410)，含复试线、综合热度、可报专业与可靠度标注。✅已核实为官方/研招网核对，🔍参考为机构整理或待核，报考前以官方目录为准。</p>"
    + mine_all
)

final_parts = [(db_parts_html[0][0], part1_content)] + db_parts_html[1:]

body_sections = []
for i, (title, html_body) in enumerate(final_parts, start=1):
    body_sections.append(f'<section id="part{i}">{html_body}</section>')

html_doc = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>408 考研终极版 · 院校大全 + 备考指南</title>
{CSS}
</head>
<body>
<header>
  <div class="wrap">
    <h1>408 考研终极版 · 院校大全 + 备考指南</h1>
    <p>综合三份资料：川渝院校专业级数据 + 全国双非精细数据 + 全套备考笔记，含专业代码、复试线、热度与可靠度标注</p>
    <span class="pill">更新：2026-08-13</span>
    <span class="pill">408=计算机学科专业基础综合(统考)</span>
    <span class="pill">11408=英一+数一+408</span>
    <span class="pill">22408=英二+数二+408</span>
  </div>
</header>
<nav class="toc"><div class="wrap">{''.join(nav_items)}</div></nav>
<div class="wrap">
{''.join(body_sections)}
<footer>
  综合整理自：豆包 408 全套资料、trae 川渝院校对比、帅帅全国双非数据。院校数据请以目标院校当年官方招生目录为准，不构成报考建议。
</footer>
</div>
</body>
</html>
"""

OUT_HTML = "408考研终极版_院校大全+备考指南_20260813.html"
with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(html_doc)
print("已生成:", OUT_HTML, "大小:", len(html_doc))


# ---------------------------------------------------------------------------
# 六、生成终极 xlsx（合并三份院校数据）
# ---------------------------------------------------------------------------
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_XLSX = "408考研终极版_院校数据_20260813.xlsx"
out_wb = openpyxl.Workbook()
out_wb.remove(out_wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def copy_sheet(src_ws, dst_name):
    dst = out_wb.create_sheet(dst_name)
    for row in src_ws.iter_rows(values_only=True):
        dst.append(list(row))
    ncols = dst.max_column
    for c in range(1, ncols + 1):
        cell = dst.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    dst.freeze_panes = "A2"
    return dst


# 我的双非数据（保留关键 sheet）
mine_sheet_map = {
    "408专业速查": "408专业速查",
    "总览": "全国双非·总览",
    "热度榜": "全国双非·热度榜",
    "重点院校": "全国双非·重点院校",
    "调剂院校": "全国双非·调剂",
    "数据源": "数据源",
}
for src_name, dst_name in mine_sheet_map.items():
    copy_sheet(mine_wb[src_name], dst_name)

# trae 川渝数据
for title, _ in trae_sheets:
    copy_sheet(trae_wb[title], "川渝·" + title)

out_wb.save(OUT_XLSX)
print("已生成:", OUT_XLSX, "sheets:", out_wb.sheetnames)
