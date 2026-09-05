# -*- coding: utf-8 -*-
"""终极版 XLSX（20260826 三源整合版）：基于新版 HTML 生成 12 个 Sheet"""
import io, json, re, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HTML = r"06_终极版输出\全国408_085410双非热度版_终极版_20260826.html"
OUT = r"06_终极版输出\全国408_085410双非热度版_终极版_20260826.xlsx"
MAIN_MD = r"01_择校与规划\085410_22408_择校与规划_20260820.md"

s = io.open(HTML, encoding="utf-8").read()

def grab(begin, end):
    a = s.index(begin) + len(begin)
    b = s.index(end, a) + 1
    return s[a:b]

S = json.loads(grab("var S=", "];\n// ===== 985/211"))
C = json.loads(grab("var C=", "];\n// ===== 初试科目映射"))
ka = s.index("var K={") + len("var K={") - 1
kb = s.index("};\nfunction exKey") + 1
K = json.loads(s[ka:kb])
SRCS = json.loads(grab("var SRCS=", "];\n\n// ===== 工具函数"))

def exKey(d):
    return K.get(d["n"] + "|" + d["c"]) or K.get(d["n"]) or "待确认"

ZONE_B = ["内蒙古", "广西", "海南", "贵州", "云南", "西藏", "甘肃", "青海", "宁夏", "新疆"]
def zoneOf(p): return "B" if p in ZONE_B else "A"
def baseLine(p): return 254 if zoneOf(p) == "B" else 264
def gapOf(d): return None if d["l"] is None else d["l"] - baseLine(d["p"])
def clamp(v): return max(0, min(100, v))
def lineScore(d):
    if d["l"] is None: return 45
    return clamp(round(gapOf(d) * 100 / 96))
def ratioScore(d):
    return 30 if d["ratio"] is None else clamp(round((d["ratio"] - 1) / 8 * 100))
def rrScore(d):
    return 35 if d["rr"] is None else clamp(round((d["rr"] - 1) / 1.5 * 100))
def fcScore(d):
    f = str(d["fc"])
    if any(x in f for x in ("大量调剂", "调剂", "一志愿不足", "缺口")): return 12
    if "一志愿为主" in f: return 72
    if f in ("待确认", "—"): return 45
    return 55
def compScore(d): return round(0.45 * lineScore(d) + 0.25 * ratioScore(d) + 0.2 * rrScore(d) + 0.1 * fcScore(d))
def heatScore(d): return round(0.7 * compScore(d) + 0.3 * (d["net"] or 0))
def gapStr(d):
    g = gapOf(d)
    return "—" if g is None else ("=国家线" if g == 0 else "+" + str(g))
def na(v): return "—" if v is None else v
def lh_str(d):
    a = list(d.get("lh") or [None, None, None]) + [d["l"]]
    return " / ".join("—" if v is None else str(v) for v in a)

def nn_cols(d):
    nn_ = d.get("nn")
    if not nn_:
        return ("—", "—", "—")
    a408 = na(nn_.get("a408"))
    lr = "—" if nn_.get("lr") is None else (str(round(nn_["lr"] * 100)) + "%")
    wd = d.get("wd")
    wd_txt = "—"
    if wd:
        wd_txt = wd["y"] + " https://…" if len(wd["u"]) > 90 else wd["y"] + " " + wd["u"]
    return (a408, lr, wd_txt)

wb = openpyxl.Workbook()
HDR_FILL = PatternFill("solid", fgColor="0F766E")
HDR_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="C8D6D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

def sheet(ws, headers, rows, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if isinstance(v, str) and v.startswith("="):
                cell.data_type = "s"  # 防 "=国家线" 被当作公式
            cell.font = BODY; cell.border = BORDER; cell.alignment = WRAP
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# 1 总览（加 408均分/录取率/王道考情 3 列，插在 综合热度 与 备注 之间）
rows = []
for d in S:
    a408, lr, wd_txt = nn_cols(d)
    rows.append([d["n"], d["p"], d["r"], d["t"], d["c"], exKey(d), d["d"], d["l"], lh_str(d), gapStr(d), d["plan"],
                 d["fc"], na(d["ratio"]), na(d["rr"]), na(d.get("rec")), na(d.get("adm")), na(d.get("max_s")),
                 na(d.get("min_s")), na(d.get("avg_s")), d.get("scope") or "—", d["ai"], d["src"], d["net"], heatScore(d),
                 a408, lr, wd_txt, d["note"]])
ws = wb.active; ws.title = "总览"
sheet(ws, ["院校","省份","大区","层次","学院","初试科目","方向","2026复试线","历年线(23/24/25/26)","线差","拟招","一志愿/调剂","报录比","复录比","复试人数","录取人数","录取最高","录取最低","录取平均","口径","AI方向备注","来源","网络热度","综合热度","408均分(N诺)","录取率(N诺)","王道考情","备注"], rows,
      [16,6,8,12,30,10,22,10,24,8,14,16,8,8,9,9,9,9,9,46,16,12,8,8,9,9,44,50])

# 2 热度榜
sortedS = sorted(S, key=heatScore, reverse=True)
rows = [[i+1, d["n"], d["r"], d["t"], d["c"], exKey(d), d["l"], gapStr(d), compScore(d), d["net"], heatScore(d), d["ai"], d["src"], d["note"]]
        for i, d in enumerate(sortedS)]
ws = wb.create_sheet("热度榜")
sheet(ws, ["排名","院校","大区","层次","学院","初试科目","2026线","线差","竞争热度","网络热度","综合热度","AI方向备注","来源","备注"], rows,
      [6,16,8,12,30,10,9,9,9,9,9,16,12,50])

# 3 区域榜
E_REG = ["华北","华东","东北"]; W_REG = ["华中","华南","西南","西北"]
rows = []
for label, regs in (("华东/华北/东北", E_REG), ("华中/华南/西南/西北", W_REG)):
    merged = [("主体", d) for d in S if d["r"] in regs] + [("对照", d) for d in C if d["r"] in regs]
    merged.sort(key=lambda x: heatScore(x[1]), reverse=True)
    for k, (cat, d) in enumerate(merged, 1):
        rows.append([label, cat, k, d["n"], d["p"], d["r"], d["t"], d["c"], exKey(d), d["l"], gapStr(d),
                     compScore(d), d["net"], heatScore(d), d["ai"], d["src"], d["note"]])
ws = wb.create_sheet("区域榜")
sheet(ws, ["榜单","类别","排名","院校","省份","大区","层次","学院","初试科目","2026线","线差","竞争热度","网络热度","综合热度","AI方向备注","来源","备注"], rows,
      [20,6,6,16,6,8,12,30,10,9,9,9,9,9,16,12,50])

# 4 招录比对比
rows = [[d["n"], d["r"], d["t"], d["c"], exKey(d), d["l"], gapStr(d), d["plan"], d["fc"], na(d["ratio"]), na(d["rr"]),
         na(d.get("rec")), na(d.get("adm")), na(d.get("max_s")), na(d.get("min_s")), na(d.get("avg_s")), compScore(d), d["src"]]
        for d in sortedS]
ws = wb.create_sheet("招录比对比")
sheet(ws, ["院校","大区","层次","学院","初试科目","2026线","线差","拟招","一志愿/调剂","报录比","复录比","复试人数","录取人数","录取最高","录取最低","录取平均","竞争热度","来源"], rows,
      [16,8,12,30,10,9,9,14,16,8,8,9,9,9,9,9,9,12])

# 4.5 考情明细
rows = []
for d in S + C:
    if any(d.get(k) is not None for k in ("rec", "adm", "max_s", "min_s", "avg_s")) or d.get("scope"):
        rows.append([d["n"], d["c"], d["t"], exKey(d), na(d["l"]), na(d["plan"]), na(d.get("rec")), na(d.get("adm")),
                     na(d.get("rr")), na(d.get("ratio")), na(d.get("max_s")), na(d.get("min_s")), na(d.get("avg_s")),
                     d.get("scope") or "—", d.get("verified") or "查无", d.get("srcu") or "—"])
ws = wb.create_sheet("考情明细")
sheet(ws, ["院校","学院","层次","初试科目","2026线","拟招","复试人数","录取人数","复录比","报录比","录取最高","录取最低","录取平均","口径","核实状态","来源URL"], rows,
      [18,30,12,12,9,14,9,9,8,8,9,9,9,60,12,70])

# 4.6 三源核对（新增）
def flag_cell(d):
    flags = []
    nn_ = d.get("nn")
    if nn_:
        if nn_["subj"] not in ("英二数二408", "推测408"):
            flags.append("N诺口径" + nn_["subj"])
        if nn_.get("line") and d.get("l") and abs(nn_["line"] - d["l"]) >= 5:
            flags.append("N诺线差" + str(nn_["line"] - d["l"]))
    cy_ = d.get("cy")
    if cy_:
        if cy_["code"] != "085410":
            flags.append("川渝参考" + cy_["code"])
        if d.get("rec") is not None and cy_.get("rc") is not None and d["rec"] != cy_["rc"]:
            flags.append("复试人数差异")
        if d.get("adm") is not None and cy_.get("ad") is not None and d["adm"] != cy_["ad"]:
            flags.append("录取人数差异")
    return "；".join(flags) if flags else "—"

rows = []
for tag, lst in (("主体", S), ("对照", C)):
    for d in lst:
        nn_ = d.get("nn")
        cy_ = d.get("cy")
        wd_ = d.get("wd")
        rows.append([d["n"], tag, d["t"], na(d.get("l")),
                     nn_["maj"] if nn_ else "—", nn_["dept"] if nn_ else "—", nn_["subj"] if nn_ else "—",
                     na(nn_.get("a408")) if nn_ else "—",
                     ("—" if nn_.get("lr") is None else str(round(nn_["lr"] * 100)) + "%") if nn_ else "—",
                     cy_["code"] if cy_ else "—",
                     na(cy_.get("rc")) if cy_ else "—", na(cy_.get("ad")) if cy_ else "—", na(cy_.get("avg")) if cy_ else "—",
                     wd_["cnt"] if wd_ else "—", wd_["years"] if wd_ else "—", wd_["u"] if wd_ else "—",
                     flag_cell(d)])
ws = wb.create_sheet("三源核对")
sheet(ws, ["院校","类别","层次","终极版2026线","N诺专业","N诺院系","N诺初试","N诺408均分","N诺录取率","川渝专业代码","川渝复试人数","川渝录取人数","川渝均分","王道链接数","王道覆盖年份","王道26链接","标注"], rows,
      [16,6,12,9,24,24,12,9,9,12,9,9,9,9,26,60,40])

# 4.7 N诺新增候选校（终极版未收录的双非408校）
import openpyxl as _ox
nn_wb = _ox.load_workbook(r"deliverables\20260826-双非408考情_N诺提取\双非408考情_2026提取.xlsx", read_only=True, data_only=True)
def nrows(ws):
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]
nn_rows = collections.defaultdict(list)
hdr = None
for r in nrows(nn_wb["2026年408系明细"]):
    if hdr is None: hdr = r; continue
    if len(r) < 19 or not r[0]: continue
    nn_rows[str(r[0]).strip()].append(r)
nn_wb.close()

def pick(rows):
    def score(r):
        maj, subj = str(r[3] or ""), str(r[4] or "").strip()
        s = 0
        if "人工智能" in maj: s += 300
        elif "智能科学" in maj or "智能技术" in maj: s += 250
        elif "电子信息" in maj: s += 200
        elif "计算机" in maj: s += 150
        elif "软件" in maj: s += 100
        if subj in ("英二数二408", "推测408"): s += 80
        elif subj == "英一数一408": s -= 40
        if isinstance(r[9], (int, float)): s += 20
        if isinstance(r[5], (int, float)): s += 10
        if isinstance(r[6], (int, float)): s += min(r[6], 800) / 100.0
        return s
    return max(rows, key=score)

def fnum(v):
    if not isinstance(v, (int, float)) or isinstance(v, bool): return None
    x = float(v)
    return int(x) if x == int(x) else round(x, 1)

def flag_big(rc, ad):
    f = []
    if rc is not None and rc > 300: f.append("复试人数待核")
    if ad is not None and ad > 250: f.append("录取人数待核")
    return "（" + "；".join(f) + "）" if f else ""

zc = set(d["n"] for d in S + C)
cand = []
for name, rows in nn_rows.items():
    if name in zc: continue
    r = pick(rows)
    maj, subj = str(r[3] or ""), str(r[4] or "").strip()
    note = []
    note.append("✅22408兼容" if subj in ("英二数二408", "推测408") else "⚠️非22408(" + subj + ")")
    if "人工智能" in maj: note.append("含AI")
    elif "电子信息" in maj: note.append("含电子信息")
    lr = fnum(r[15])
    rc = fnum(r[6]); ad = fnum(r[13])
    cand.append([name, r[1], r[2], maj, subj, fnum(r[5]), rc, fnum(r[7]), fnum(r[8]), fnum(r[9]),
                 fnum(r[10]), fnum(r[11]), fnum(r[12]), ad, fnum(r[14]),
                 ("—" if lr is None else str(round(lr * 100)) + "%"), fnum(r[17]),
                 "；".join(note) + flag_big(rc, ad)])
cand.sort(key=lambda x: (x[4] in ("英二数二408", "推测408"), "含AI" in x[17], x[15] if isinstance(x[15], str) else "0"), reverse=True)
ws = wb.create_sheet("N诺新增候选校")
sheet(ws, ["院校","省份","院系","专业","初试","复试线","复试人数","调剂人数","复试总分均分","408均分","政治均分","英语均分","数学均分","录取人数","录取均分","录取率","报录比(N诺=复试/录取)","备注"], cand,
      [16,6,24,22,12,9,9,9,11,9,9,9,9,9,9,9,11,34])

# 5 调剂院校
adj = [d for d in sortedS if any(x in str(d["fc"]) for x in ("调剂", "一志愿不足", "缺口"))]
rows = [[d["n"], exKey(d), d["r"], d["t"], d["l"], d["fc"], d["ai"], d["src"]] for d in adj]
ws = wb.create_sheet("调剂院校")
sheet(ws, ["院校","初试科目","大区","层次","2026线","一志愿/调剂","AI方向","来源"], rows,
      [16,10,8,12,9,18,16,12])

# 6 重点院校
bySchool = {}
for d in S:
    bySchool.setdefault(d["n"], []).append(d)
top = sorted(bySchool.items(), key=lambda kv: max(heatScore(x) for x in kv[1]), reverse=True)[:12]
rows = []
for name, ds in top:
    main = max(ds, key=heatScore)
    rows.append([name, main["p"], main["r"], main["t"], main["c"], exKey(main), main["d"], main["l"], main["plan"],
                 main["fc"], main["ai"], main["src"], heatScore(main), "；".join(d["note"] for d in ds)[:200]])
ws = wb.create_sheet("重点院校")
sheet(ws, ["院校","省份","大区","层次","学院","初试科目","方向","2026线","拟招","一志愿/调剂","AI方向","来源","综合热度","备注"], rows,
      [16,6,8,12,30,10,22,9,14,16,16,12,9,50])

# 7 数据源
rows = [[t, u] for t, u in SRCS]
ws = wb.create_sheet("数据源")
sheet(ws, ["来源","URL"], rows, [42, 90])

# 8 2027改考动态
md = io.open(MAIN_MD, encoding="utf-8").read()
lines = md.split("\n")
gk_rows = []
in_gk = False
for l in lines:
    if l.startswith("##") and "2027改考动态" in l:
        in_gk = True
        continue
    if in_gk:
        if l.startswith("##"):
            break
        if l.startswith("|") and not l.startswith("|---"):
            cells = [c.strip() for c in l.strip("|").split("|")]
            if cells and cells[0] not in ("院校", ""):
                gk_rows.append(cells)
ws = wb.create_sheet("2027改考动态")
sheet(ws, ["院校","层次","专业/范围","原科目","新科目","生效年份","来源","备注"], gk_rows,
      [22,10,34,16,16,10,20,50])

# 9 导师研究方向
ds_rows = []
in_ds = False
for l in lines:
    if l.startswith("##") and "导师研究方向" in l:
        in_ds = True
        continue
    if in_ds:
        if l.startswith("##"):
            break
        m = re.match(r"- \*\*(.+?)\*\*：(.+)", l)
        if m:
            ds_rows.append([m.group(1), m.group(2)])
ws = wb.create_sheet("导师研究方向")
sheet(ws, ["院校", "导师及研究方向（含Agent/大模型/具身智能关键词）"], ds_rows, [20, 110])

wb.save(OUT)
print("已写入:", OUT)
print("S:", len(S), "C:", len(C), "SRCS:", len(SRCS),
      "三源核对行:", len(rows_ck := None) if False else "", "候选校:", len(cand),
      "Sheet数:", len(wb.sheetnames))
