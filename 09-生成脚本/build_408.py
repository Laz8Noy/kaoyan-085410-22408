# -*- coding: utf-8 -*-
"""
408 考研双非院校项目 —— 增强版生成脚本
功能：
  1. 读取现有"全国408_085410双非热度版"数据
  2. 补充"专业"维度（408 可报专业速查 + 每校可报专业）
  3. 提升可靠度标注
  4. 生成增强版 xlsx 与高可读性 HTML
"""
import glob
import sys
import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# 一、408 可报专业速查（基于研招网/官方目录核实，2026 招生年度）
# ---------------------------------------------------------------------------
# (代码, 专业名称, 学位类型, 常见初试科目组合, 说明)
MAJORS = [
    ("081200", "计算机科学与技术", "学硕", "11408（英一+数一+408）", "最核心的计算机学硕，报考面广、调剂认可度高"),
    ("083500", "软件工程", "学硕", "11408（英一+数一+408）", "软件方向学硕，部分院校合并进计算机学院招生"),
    ("083900", "网络空间安全", "学硕", "11408（英一+数一+408）", "网安方向学硕，与 085412 对应"),
    ("140500", "智能科学与技术", "学硕", "11408 或自命题", "交叉学科，部分院校考 408"),
    ("085404", "计算机技术", "专硕", "22408（英二+数二+408）或11408", "计算机专硕主力，招生名额通常最多"),
    ("085405", "软件工程", "专硕", "22408（英二+数二+408）", "软件方向专硕"),
    ("085410", "人工智能", "专硕", "22408 或 11408", "本表主体专业，近年热度高、扩招明显"),
    ("085411", "大数据技术与工程", "专硕", "22408（英二+数二+408）", "大数据方向专硕"),
    ("085412", "网络与信息安全", "专硕", "22408（英二+数二+408）", "网安方向专硕，与 083900 对应"),
    ("085401", "新一代电子信息技术", "专硕", "22408 或自命题", "电子信息大类下方向，部分院校考 408"),
]

COMBOS = {
    "11408": "政治 + 英语一 + 数学一 + 408",
    "22408": "政治 + 英语二 + 数学二 + 408",
}

# ---------------------------------------------------------------------------
# 二、重点院校可报 408 专业（已逐校联网核实）
# ---------------------------------------------------------------------------
SCHOOL_MAJORS = {
    "深圳大学": "081200计算机科学与技术、085404计算机技术、085410人工智能、085411大数据（计算机类学院考408）",
    "重庆邮电大学": "081200计算机科学与技术、083500软件工程、085404计算机技术、085405软件工程、085410人工智能",
    "西安邮电大学": "081200计算机科学与技术、083500软件工程、085404计算机技术、085411大数据、085410人工智能",
    "湖南科技大学": "081200计算机科学与技术、083500软件工程、085404计算机技术、085405软件工程、085410人工智能",
    "广州大学": "081200计算机科学与技术、085404计算机技术、085405软件工程、085410人工智能、085412网络与信息安全",
    "广东工业大学": "081200计算机科学与技术、083500软件工程、085404计算机技术、085405软件工程、085410人工智能、085412网络与信息安全",
    "杭州电子科技大学": "081200计算机科学与技术、085404计算机技术、085405软件工程（计算机学院考408）；085410人工智能（初试861非408）",
    "浙江理工大学": "085404计算机技术、085410人工智能（22408）",
}


def infer_majors(school, college):
    """根据学院名推断该校主要 408 专业（非逐校核实的部分，标注为参考）。"""
    if school in SCHOOL_MAJORS:
        return SCHOOL_MAJORS[school]
    c = college or ""
    if "网络空间安全" in c or "网安" in c:
        return "085410人工智能 + 083900网络空间安全/085412网络与信息安全（以官方目录为准）"
    if "计算机" in c:
        return "085410人工智能 + 081200计算机科学与技术/085404计算机技术（以官方目录为准）"
    if "人工智能" in c or "智能科学" in c or "智能与" in c:
        return "085410人工智能 + 可能081200计算机（以官方目录为准）"
    return "085410人工智能（本表主体，学院归属见备注；以官方目录为准）"


# ---------------------------------------------------------------------------
# 三、读取现有数据
# ---------------------------------------------------------------------------
SRC = glob.glob("*408_085410*.xlsx")[0]
wb_src = openpyxl.load_workbook(SRC, data_only=True)

def read_sheet(name):
    ws = wb_src[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows

overview = read_sheet("总览")
area = read_sheet("区域榜")
recruit = read_sheet("招录比对比")
adjust = read_sheet("调剂院校")
focus = read_sheet("重点院校")
datasrc = read_sheet("数据源")

# 总览表头
ov_header = overview[0]
# 找到"初试科目"和"学院"列索引，用于补专业
def col_index(header, name):
    for i, h in enumerate(header):
        if h and name in str(h):
            return i
    return -1

i_college = col_index(ov_header, "学院")
i_school = col_index(ov_header, "院校")

# 总览数据行（跳过表头）
ov_rows = [r for r in overview[1:] if r and r[0]]


# ---------------------------------------------------------------------------
# 四、生成增强版 xlsx
# ---------------------------------------------------------------------------
OUT_XLSX = "全国408_双非院校_完整版_20260813.xlsx"

wb = openpyxl.Workbook()

# 样式
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"

def autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 1: 专业速查 ---
ws1 = wb.active
ws1.title = "408专业速查"
ws1.append(["专业代码", "专业名称", "学位类型", "常见初试科目", "说明"])
for m in MAJORS:
    ws1.append(list(m))
style_header(ws1, 5)
autowidth(ws1, [10, 18, 9, 26, 44])
for r in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=5):
    for cell in r:
        cell.border = BORDER
        cell.alignment = CENTER if cell.column <= 4 else LEFT

# --- Sheet 2: 总览（增强：新增"可报专业"与"专业可靠度"列）---
ws2 = wb.create_sheet("总览")
new_header = list(ov_header) + ["可报专业(408)", "专业可靠度"]
ws2.append(new_header)
for row in ov_rows:
    school = str(row[i_school]) if i_school >= 0 else ""
    college = str(row[i_college]) if i_college >= 0 else ""
    majors = infer_majors(school, college)
    rel = "✅已核实" if school in SCHOOL_MAJORS else "🔍参考"
    ws2.append(row + [majors, rel])
style_header(ws2, len(new_header))
autowidth(ws2, [16, 7, 6, 8, 22, 8, 14, 8, 6, 7, 12, 7, 7, 9, 7, 7, 8, 22, 30, 9])
for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(new_header)):
    for cell in r:
        cell.border = BORDER
        cell.alignment = LEFT if cell.column >= 18 else CENTER

# --- Sheet 3: 热度榜 ---
ws3 = wb.create_sheet("热度榜")
hot = read_sheet("热度榜")
hot_header = hot[0] + ["可报专业(408)", "专业可靠度"]
ws3.append(hot_header)
for row in hot[1:]:
    if not row or not row[0]:
        continue
    school = str(row[1]) if len(row) > 1 else ""
    college = str(row[4]) if len(row) > 4 else ""
    ws3.append(row + [infer_majors(school, college), "✅已核实" if school in SCHOOL_MAJORS else "🔍参考"])
style_header(ws3, len(hot_header))
autowidth(ws3, [6, 16, 7, 8, 22, 8, 8, 6, 8, 8, 8, 9, 7, 30, 30, 9])
for r in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(hot_header)):
    for cell in r:
        cell.border = BORDER
        cell.alignment = LEFT if cell.column >= 14 else CENTER

# --- Sheet 4: 重点院校 ---
ws4 = wb.create_sheet("重点院校")
focus = read_sheet("重点院校")
focus_header = focus[0] + ["可报专业(408)", "专业可靠度"]
ws4.append(focus_header)
for row in focus[1:]:
    if not row or not row[0]:
        continue
    school = str(row[0])
    college = str(row[4]) if len(row) > 4 else ""
    ws4.append(row + [infer_majors(school, college), "✅已核实" if school in SCHOOL_MAJORS else "🔍参考"])
style_header(ws4, len(focus_header))
autowidth(ws4, [16, 7, 6, 8, 24, 8, 14, 8, 7, 11, 9, 7, 8, 30, 30, 9])
for r in ws4.iter_rows(min_row=2, max_row=ws4.max_row, max_col=len(focus_header)):
    for cell in r:
        cell.border = BORDER
        cell.alignment = LEFT if cell.column >= 14 else CENTER

# --- Sheet 5: 招录比对比（保留） ---
ws5 = wb.create_sheet("招录比对比")
for row in recruit:
    ws5.append(row)
style_header(ws5, len(recruit[0]))
autowidth(ws5, [16, 7, 8, 22, 8, 8, 6, 7, 12, 7, 7, 8, 7])
for r in ws5.iter_rows(min_row=2, max_row=ws5.max_row, max_col=len(recruit[0])):
    for cell in r:
        cell.border = BORDER
        cell.alignment = CENTER

# --- Sheet 6: 调剂院校 ---
ws6 = wb.create_sheet("调剂院校")
for row in adjust:
    ws6.append(row)
style_header(ws6, len(adjust[0]))
autowidth(ws6, [16, 8, 7, 8, 8, 12, 9, 7])
for r in ws6.iter_rows(min_row=2, max_row=ws6.max_row, max_col=len(adjust[0])):
    for cell in r:
        cell.border = BORDER
        cell.alignment = CENTER

# --- Sheet 7: 区域榜 ---
ws7 = wb.create_sheet("区域榜")
for row in area:
    ws7.append(row)
style_header(ws7, len(area[0]))
autowidth(ws7, [14, 12, 6, 16, 7, 6, 8, 22, 8, 8, 6, 8, 8, 8, 9, 7, 22])
for r in ws7.iter_rows(min_row=2, max_row=ws7.max_row, max_col=len(area[0])):
    for cell in r:
        cell.border = BORDER
        cell.alignment = CENTER

# --- Sheet 8: 数据源 ---
ws8 = wb.create_sheet("数据源")
for row in datasrc:
    ws8.append(row)
style_header(ws8, 2)
autowidth(ws8, [36, 80])
for r in ws8.iter_rows(min_row=2, max_row=ws8.max_row, max_col=2):
    for cell in r:
        cell.alignment = LEFT

wb.save(OUT_XLSX)
print("已生成:", OUT_XLSX)


# ---------------------------------------------------------------------------
# 五、生成高可读性 HTML
# ---------------------------------------------------------------------------
OUT_HTML = "全国408_双非院校_完整版_20260813.html"


def esc(s):
    if s is None:
        return ""
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def heat_style(score, low=20, high=70):
    """根据综合热度返回背景色，热度越高越暖。始终保留文字数字。"""
    try:
        v = float(score)
    except (TypeError, ValueError):
        return ""
    t = max(0.0, min(1.0, (v - low) / (high - low)))
    # 浅蓝 -> 浅黄 -> 橙 -> 红
    if t < 0.33:
        r, g, b = 235, 245, 255
    elif t < 0.66:
        r, g, b = 255, 242, 204
    elif t < 0.85:
        r, g, b = 255, 214, 153
    else:
        r, g, b = 255, 170, 150
    return f'style="background:rgb({r},{g},{b});font-weight:600"'


def line_style(line):
    """复试线颜色：高于国家线较多标暖色，接近国家线标冷色。"""
    try:
        v = int(line)
    except (TypeError, ValueError):
        return ""
    if v >= 320:
        return 'style="color:#c0392b;font-weight:700"'
    if v >= 280:
        return 'style="color:#e67e22;font-weight:600"'
    return 'style="color:#1e8449;font-weight:600"'


# 专业速查表 HTML
major_rows = []
for code, name, degree, subj, desc in MAJORS:
    tag_cls = "deg-xue" if degree == "学硕" else "deg-zhuan"
    major_rows.append(
        f'<tr><td class="code">{esc(code)}</td>'
        f'<td class="mname">{esc(name)}</td>'
        f'<td><span class="tag {tag_cls}">{esc(degree)}</span></td>'
        f'<td class="subj">{esc(subj)}</td>'
        f'<td class="desc">{esc(desc)}</td></tr>'
    )

# 热度榜 TOP 15
hot_rows = []
for row in hot[1:]:
    if not row or not row[0]:
        continue
    rank, school, area_, lvl, college, subj, line, diff = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
    comp, net, total, ai, src, note = row[8], row[9], row[10], row[11], row[12], row[13]
    school_txt = esc(school)
    majors = infer_majors(str(school), str(college))
    rel = "已核实" if str(school) in SCHOOL_MAJORS else "参考"
    rel_cls = "rel-ok" if rel == "已核实" else "rel-ref"
    hot_rows.append(
        f'<tr>'
        f'<td class="rank">{esc(rank)}</td>'
        f'<td class="school">{school_txt}<div class="majors">{esc(majors)}</div></td>'
        f'<td>{esc(area_)}</td>'
        f'<td>{esc(college)}</td>'
        f'<td>{esc(subj)}</td>'
        f'<td {line_style(line)}>{esc(line)}</td>'
        f'<td>{esc(diff)}</td>'
        f'<td {heat_style(total)}>{esc(total)}</td>'
        f'<td><span class="tag {rel_cls}">{rel}</span></td>'
        f'<td class="note">{esc(note)}</td>'
        f'</tr>'
    )

# 总览完整表
ov_html_rows = []
for row in ov_rows:
    school = str(row[i_school]) if i_school >= 0 else ""
    college = str(row[i_college]) if i_college >= 0 else ""
    majors = infer_majors(school, college)
    rel = "已核实" if school in SCHOOL_MAJORS else "参考"
    rel_cls = "rel-ok" if rel == "已核实" else "rel-ref"
    prov, area_, lvl, subj = row[1], row[2], row[3], row[5]
    line, diff, plan, fs = row[7], row[8], row[9], row[10]
    total, src, note = row[16], row[14], row[17]
    ov_html_rows.append(
        f'<tr data-search="{esc(school)}{esc(college)}{esc(majors)}">'
        f'<td class="school">{esc(school)}</td>'
        f'<td>{esc(prov)}</td>'
        f'<td>{esc(area_)}</td>'
        f'<td>{esc(lvl)}</td>'
        f'<td>{esc(college)}</td>'
        f'<td>{esc(subj)}</td>'
        f'<td {line_style(line)}>{esc(line)}</td>'
        f'<td>{esc(diff)}</td>'
        f'<td>{esc(plan)}</td>'
        f'<td>{esc(fs)}</td>'
        f'<td {heat_style(total)}>{esc(total)}</td>'
        f'<td class="majors">{esc(majors)}</td>'
        f'<td><span class="tag {rel_cls}">{rel}</span></td>'
        f'<td class="note">{esc(note)}</td>'
        f'</tr>'
    )

# 重点院校详情卡片
focus_cards = []
for row in focus[1:]:
    if not row or not row[0]:
        continue
    school, prov, area_, lvl, college = row[0], row[1], row[2], row[3], row[4]
    subj, line, plan, ai, total, note = row[5], row[7], row[8], row[10], row[12], row[13]
    majors = infer_majors(str(school), str(college))
    rel = "已核实" if str(school) in SCHOOL_MAJORS else "参考"
    rel_cls = "rel-ok" if rel == "已核实" else "rel-ref"
    focus_cards.append(
        f'<div class="card">'
        f'<div class="card-head"><span class="school">{esc(school)}</span>'
        f'<span class="tag {rel_cls}">{rel}</span></div>'
        f'<div class="card-meta">{esc(prov)} · {esc(area_)} · {esc(lvl)} · 复试线 <b {line_style(line)}>{esc(line)}</b> · 热度 {esc(total)}</div>'
        f'<div class="card-college">学院：{esc(college)}　初试：{esc(subj)}　拟招：{esc(plan)}　AI：{esc(ai)}</div>'
        f'<div class="majors">可报专业：{esc(majors)}</div>'
        f'<div class="note">{esc(note)}</div>'
        f'</div>'
    )

# 调剂院校
adj_rows = []
for row in adjust[1:]:
    if not row or not row[0]:
        continue
    school, subj, area_, lvl, line, fs, ai, src = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
    adj_rows.append(
        f'<tr><td>{esc(school)}</td><td>{esc(area_)}</td><td>{esc(subj)}</td>'
        f'<td>{esc(line)}</td><td>{esc(fs)}</td><td>{esc(ai)}</td></tr>'
    )

today = "2026-08-13"

html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>408 考研双非院校 · 完整版（专业+热度+可靠度）</title>
<style>
:root {{
  --blue:#1F4E78; --blue2:#2E75B6; --bg:#f4f6f9; --card:#ffffff;
  --line:#e3e8ef; --text:#1f2933; --muted:#6b7280; --ok:#1e8449; --ref:#b7791f;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:-apple-system,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;
  background:var(--bg); color:var(--text); line-height:1.6; }}
.wrap {{ max-width:1180px; margin:0 auto; padding:16px; }}
header {{ background:linear-gradient(135deg,#1F4E78,#2E75B6); color:#fff; padding:28px 18px; }}
header h1 {{ margin:0 0 6px; font-size:24px; }}
header p {{ margin:0; opacity:.9; font-size:14px; }}
.pill {{ display:inline-block; background:rgba(255,255,255,.18); border-radius:14px;
  padding:2px 10px; font-size:12px; margin:8px 6px 0 0; }}
section {{ margin-top:18px; }}
h2 {{ font-size:19px; margin:0 0 10px; padding-left:10px; border-left:4px solid var(--blue2); }}
.note-block {{ background:#fff8e6; border:1px solid #f0d9a0; border-radius:10px; padding:12px 14px;
  font-size:14px; color:#7a5b13; }}
.legend {{ display:flex; flex-wrap:wrap; gap:8px; margin:8px 0; font-size:13px; color:var(--muted); }}
.legend span {{ display:inline-flex; align-items:center; gap:5px; }}
.dot {{ width:12px; height:12px; border-radius:3px; display:inline-block; }}
table {{ width:100%; border-collapse:collapse; background:var(--card); font-size:13px;
  border-radius:10px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
th {{ background:var(--blue); color:#fff; padding:9px 7px; text-align:center; font-weight:600;
  white-space:nowrap; position:sticky; top:0; }}
td {{ padding:8px 7px; border-bottom:1px solid var(--line); text-align:center; vertical-align:top; }}
tr:nth-child(even) td {{ background:#f8fafc; }}
tbody tr:hover td {{ background:#eef4fb; }}
.school {{ font-weight:600; text-align:left; white-space:nowrap; }}
.majors {{ font-size:12px; color:var(--muted); text-align:left; white-space:normal; }}
.note {{ text-align:left; color:var(--muted); font-size:12px; }}
.code {{ font-family:ui-monospace,Consolas,monospace; font-weight:700; color:var(--blue2); }}
.tag {{ display:inline-block; border-radius:10px; padding:1px 8px; font-size:11px; white-space:nowrap; }}
.deg-xue {{ background:#e8f0fe; color:#1a56a8; }}
.deg-zhuan {{ background:#e6f7ee; color:#16794a; }}
.rel-ok {{ background:#e6f7ee; color:#16794a; }}
.rel-ref {{ background:#fef3e2; color:#a8641a; }}
.rank {{ font-weight:700; color:var(--blue2); }}
.cards {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(320px,1fr)); gap:12px; }}
.card {{ background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px;
  box-shadow:0 1px 3px rgba(0,0,0,.06); }}
.card-head {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:6px; }}
.card-head .school {{ font-size:16px; }}
.card-meta {{ font-size:13px; color:var(--muted); margin:4px 0; }}
.card-college {{ font-size:13px; margin:4px 0; }}
.card .note {{ font-size:12px; }}
.search {{ width:100%; padding:10px 12px; border:1px solid var(--line); border-radius:10px;
  font-size:14px; margin-bottom:10px; }}
.table-scroll {{ overflow-x:auto; border-radius:10px; }}
.grid2 {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; }}
.info-card {{ background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px; }}
.info-card h3 {{ margin:0 0 8px; font-size:15px; color:var(--blue); }}
.info-card ul {{ margin:6px 0 0; padding-left:18px; font-size:13px; }}
.info-card li {{ margin:4px 0; }}
footer {{ text-align:center; color:var(--muted); font-size:12px; margin:28px 0 10px; }}
@media (max-width:720px) {{
  header h1 {{ font-size:20px; }}
  .grid2 {{ grid-template-columns:1fr; }}
  .cards {{ grid-template-columns:1fr; }}
  th,td {{ font-size:12px; padding:6px 4px; }}
}}
</style>
</head>
<body>
<header>
  <div class="wrap">
    <h1>408 考研双非院校 · 完整版</h1>
    <p>专业代码 + 热度榜 + 复试线 + 可靠度标注，聚焦「考 408 的双非院校」</p>
    <span class="pill">更新：{today}</span>
    <span class="pill">408 = 计算机学科专业基础综合（统考）</span>
    <span class="pill">11408 = 英一+数一+408</span>
    <span class="pill">22408 = 英二+数二+408</span>
  </div>
</header>

<div class="wrap">

  <div class="note-block">
    <b>先搞清楚两件事：</b>① <b>11408</b> = 政治 + 英语一 + 数学一 + 408（学硕标准，数学含概率论，难度高）；
    ② <b>22408</b> = 政治 + 英语二 + 数学二 + 408（专硕常见，数二不含概率论，复习量少约 1/3）。
    近年越来越多专硕也改成 11408，报考前务必以<b>目标院校当年招生目录</b>为准。
  </div>

  <section>
    <h2>① 408 能报哪些专业（专业速查）</h2>
    <div class="legend">
      <span><i class="dot" style="background:#e8f0fe"></i>学硕（学术学位）</span>
      <span><i class="dot" style="background:#e6f7ee"></i>专硕（专业学位）</span>
    </div>
    <div class="table-scroll">
      <table>
        <thead><tr><th>专业代码</th><th>专业名称</th><th>学位</th><th>常见初试科目</th><th>说明</th></tr></thead>
        <tbody>{''.join(major_rows)}</tbody>
      </table>
    </div>
    <p class="legend" style="margin-top:8px">补充：081100 控制科学与工程、081000 信息与通信工程、083700 安全科学与工程、145200 密码等专业在部分院校也考 408，需按校查目录。</p>
  </section>

  <section>
    <h2>② 综合热度榜（TOP，按热度+复试线）</h2>
    <div class="legend">
      <span><i class="dot" style="background:#eef5ff"></i>热度较低</span>
      <span><i class="dot" style="background:#fff2cc"></i>热度中</span>
      <span><i class="dot" style="background:#ffd699"></i>热度较高</span>
      <span><i class="dot" style="background:#ffaa96"></i>热度高</span>
      <span style="margin-left:12px"><i class="dot" style="background:#e6f7ee"></i>✅已核实</span>
      <span><i class="dot" style="background:#fef3e2"></i>🔍参考</span>
    </div>
    <div class="table-scroll">
      <table>
        <thead><tr><th>排名</th><th>院校</th><th>大区</th><th>学院</th><th>初试</th><th>2026线</th><th>线差</th><th>热度</th><th>可靠度</th><th>备注</th></tr></thead>
        <tbody>{''.join(hot_rows)}</tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>③ 重点院校详情</h2>
    <div class="cards">{''.join(focus_cards)}</div>
  </section>

  <section>
    <h2>④ 完整院校总览（可搜索）</h2>
    <input class="search" id="q" type="text" placeholder="搜索院校 / 学院 / 专业，例如：深圳大学、081200、人工智能…">
    <div class="table-scroll">
      <table id="all">
        <thead><tr><th>院校</th><th>省</th><th>大区</th><th>层次</th><th>学院</th><th>初试</th><th>2026线</th><th>线差</th><th>拟招</th><th>一志愿/调剂</th><th>热度</th><th>可报专业(408)</th><th>可靠度</th><th>备注</th></tr></thead>
        <tbody>{''.join(ov_html_rows)}</tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>⑤ 调剂友好院校</h2>
    <div class="table-scroll">
      <table>
        <thead><tr><th>院校</th><th>大区</th><th>初试</th><th>2026线</th><th>一志愿/调剂</th><th>AI方向</th></tr></thead>
        <tbody>{''.join(adj_rows)}</tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>⑥ 数据可靠度说明</h2>
    <div class="grid2">
      <div class="info-card">
        <h3>✅ 已核实（官方/研招网）</h3>
        <ul>
          <li>复试线、招生专业来自院校研究生院或研招网公示。</li>
          <li>重点院校（深大、重邮、西邮、湖科大、广工、广大、杭电、浙理工）的可报专业已逐校核对。</li>
        </ul>
      </div>
      <div class="info-card">
        <h3>🔍 参考（机构整理/待核）</h3>
        <ul>
          <li>标注"机构估算""待确认"的数据，建议报考前到目标院校官网二次确认。</li>
          <li>"可报专业"基于学院设置合理列出，最终以当年招生目录为准。</li>
        </ul>
      </div>
    </div>
  </section>

  <footer>
    数据采集自院校官网、研招网及公开考研信息平台，供择校参考，不构成报考建议。
    <br>完整数据源见同名 Excel 的「数据源」表。
  </footer>
</div>

<script>
document.getElementById('q').addEventListener('input', function(){{
  var kw = this.value.trim().toLowerCase();
  var rows = document.querySelectorAll('#all tbody tr');
  rows.forEach(function(r){{
    r.style.display = (r.getAttribute('data-search') || '').toLowerCase().indexOf(kw) >= 0 ? '' : 'none';
  }});
}});
</script>
</body>
</html>
"""

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(html)
print("已生成:", OUT_HTML)
